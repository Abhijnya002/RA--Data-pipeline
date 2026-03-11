"""
Section 301 List 2 (USTR-2018-0032) - Extract data from the standard exclusion request form PDF.
Output matches Complete-11k-id-attributes schema (same columns as data-extraction + add_date_col + add_withdrawn pipeline).
Filenames vary - no pattern. We download ALL PDF attachments, read each one, and identify
the standard form by its content. Uses free LLMs (Groq/Ollama) for better form identification
and OCR extraction when available.

Setup for LLM:
  - Local Llama (default): Run `ollama run llama3` - no API key, no rate limits
  - Vision (form identification): Run `ollama run llava` for image-based form scoring
  - Groq (fallback): Set GROQ_API_KEY env var for cloud fallback
  - Scanned PDFs: Marked as "Scanned" in Notes column, highlighted red (extraction not attempted)
  - Primex logo detection: pip install opencv-python (optional, for blue Primex logo on 1st page)
"""
import pandas as pd
import requests
import fitz
import os
import re
import time
import io
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import ImageEnhance, ImageOps
    HAS_OCR = True  # Optional: pip install pytesseract pdf2image; tesseract-ocr system package
except ImportError:
    HAS_OCR = False

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import cv2
    import numpy as np
    HAS_CV2 = True
except ImportError:
    HAS_CV2 = False

import section301_detection  # GitHub data-extraction.py logic - main and only PDF form detection

# Primex logo reference for image-based detection (blue stylized P)
PRIMEX_LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "primex.png")

# Config
DOCKET_ID = "USTR-2018-0032"
API_KEY =  "rWRXjXvflW9STOTgc8JKH9kFkkyY7j5qR3GcWba6" # Register at api.data.gov for higher limits
MAX_WORKERS = 8
USE_LLM = True   # Use LLM for form identification + OCR extraction (better accuracy)
PREFER_LOCAL_LLM = True   # True = use local Ollama (Llama) first; False = Groq first
OLLAMA_URL = "http://localhost:11434"  # Local Llama via Ollama (run: ollama run llama3)
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "llama3")  # e.g. llama3, llama3.2, llama3.1
OLLAMA_VISION_MODEL = os.environ.get("OLLAMA_VISION_MODEL", "llava")  # For faint/blank pages: llava, llava:13b
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")   # Fallback: https://console.groq.com
START_INDEX = 0   # For batch: process doc_ids[START_INDEX:END_INDEX]
END_INDEX = None  # None = process all
LIMIT = None      # Max docs to process (for testing), None = no limit
SKIP_USTR_RESPONSE_DATE = False  # Fetch Re_ attachment for USTR Response Date
FAST_MODE = False                # Use vision LLM + text LLM for scoring (379 rows to check - worth the accuracy)
# Path to existing Excel - rows with extraction_method="Form Fields" are skipped (kept as-is), rest are re-processed
EXISTING_EXCEL_PATH = "section301_list2_output/section301_list2_20260306_215829_UPDATED_REPROCESSED_WITHDRAWN_FILLED_REPROCESSED_REPROCESSED.xlsx"
FORM_SCORE_THRESHOLD = 8  # GitHub detection threshold (section301_detection.DETECTION_SCORE_THRESHOLD)
DEBUG_DIR = "debug_extractions_ustr"
OUTPUT_DIR = "section301_list2_output"
os.makedirs(DEBUG_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# --- LLM helpers (local Ollama / Groq) ---
def _llm_call(messages, max_tokens=1500):
    """Call LLM: local Ollama (Llama) first if PREFER_LOCAL_LLM, else Groq first."""
    def try_ollama():
        try:
            r = requests.post(f"{OLLAMA_URL}/api/chat",
                json={"model": OLLAMA_MODEL, "messages": messages, "stream": False}, timeout=120)
            if r.status_code == 200:
                return r.json().get("message", {}).get("content", "").strip()
        except Exception:
            pass
        return None
    def try_groq():
        if not GROQ_API_KEY:
            return None
        try:
            r = requests.post("https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
                json={"model": "llama-3.3-70b-versatile", "messages": messages, "max_tokens": max_tokens},
                timeout=60)
            if r.status_code == 200:
                return r.json()["choices"][0]["message"]["content"].strip()
        except Exception:
            pass
        return None
    if PREFER_LOCAL_LLM:
        out = try_ollama() or try_groq()
    else:
        out = try_groq() or try_ollama()
    return out

def llm_identify_standard_form(text_sample):
    """Use LLM to determine if PDF is the standard Section 301 exclusion request form. Returns score 0-100."""
    if not USE_LLM or len(text_sample) < 50:
        return None
    sample = text_sample[:3000]  # Limit tokens
    prompt = """You are analyzing PDF content from U.S. Section 301 China tariff exclusion requests.
The STANDARD FORM has: Requestor Information, 10-digit HTSUS code, product description, relationship to product, 
domestic/global availability, 2015/2016/2017 value/quantity, BCI status, certification. It is a structured form.
Other PDFs may be: cover letters, BCI versions, rebuttals, supporting docs, letters - NOT the main exclusion form.

Does this content appear to be the STANDARD exclusion request form (the main form firms fill out)?
Reply with ONLY a number 0-100: 100=definitely the form, 0=definitely not. No explanation.

Content:
"""
    out = _llm_call([{"role": "user", "content": prompt + sample}], max_tokens=20)
    if out:
        m = re.search(r'\b(\d{1,3})\b', out)
        if m:
            return min(100, max(0, int(m.group(1))))
    return None

def llm_extract_fields(text):
    """Use LLM to extract structured fields from OCR/text. Returns dict."""
    if not USE_LLM or len(text) < 30:
        return {}
    sample = text[:6000]
    prompt = """Extract these fields from this Section 301 exclusion request form text. Reply as JSON only.
Fields: q1_bci_status, q2_product_description, q3_htsus, q4_requestor_name, q4_organization, q4_representative,
q5_relationship, q6_attachments, q7_attachment_bci, q8_us_sources, q9_third_countries,
q10_2015_value, q10_2016_value, q10_2017_value, q11_supporting_info.
Use empty string "" for missing. For yes/no use YES, NO, or N/A.

Text:
"""
    out = _llm_call([{"role": "user", "content": prompt + sample}], max_tokens=800)
    if out:
        try:
            json_str = re.search(r'\{[\s\S]*\}', out)
            if json_str:
                return {k: str(v).strip() for k, v in __import__("json").loads(json_str.group(0)).items()}
        except Exception:
            pass
    return {}

VISION_IDENTIFY_PROMPT = """Look at this document image. Is it a U.S. Section 301 China tariff exclusion request form? The standard form has this structure:
- Question 1: BCI status (Public Document / Public Version of BCI / BCI)
- Question 2: Product description
- Question 3: 10-digit HTSUS code
- Question 4: Requestor Name, Organization Name, Representative
- Question 5: Relationship to product
- Questions 6-7: Attachments, BCI for attachments
- Questions 8-9: Domestic availability, Global availability
- Question 10: 2015/2016/2017 values
- Question 11: Supporting info

Reply with ONLY a number 0-100: 100=definitely this form (same template), 0=not this form. No explanation."""

def _is_mostly_blank_image(img):
    """Detect nearly blank pages (white + scan noise) - vision LLM hallucinates on these.
    Relaxed: faint vertical lines, light bands (like pattern PDFs) are NOT considered blank."""
    try:
        gray = img.convert("L")
        arr = list(gray.getdata())
        total = len(arr)
        dark = sum(1 for p in arr if p < 230)  # non-white pixels (relaxed from 220)
        # Only skip when < 0.1% dark (truly blank); faint pattern PDFs have structure
        if total > 0 and dark / total < 0.001:
            return True
        try:
            import statistics
            v = statistics.variance(arr)
            if v < 30:  # nearly uniform (relaxed from 100 - faint lines have some variance)
                return True
        except Exception:
            pass
        return False
    except Exception:
        return False

def _preprocess_image_for_vision(img):
    """Enhance faint/blank scans for better vision LLM readability (autocontrast, contrast boost).
    Uses old script logic: autocontrast + stronger contrast for faint pattern PDFs."""
    try:
        from PIL import ImageEnhance, ImageOps
        img = img.convert("L")
        img = ImageOps.autocontrast(img)
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(2.5)  # Stronger for faint pattern PDFs (was 2.0)
        return img.convert("RGB")
    except Exception:
        return img.convert("RGB") if img.mode != "RGB" else img

def _extract_ocr_text_for_scoring(pdf_path, max_pages=5):
    """OCR with old script preprocessing (autocontrast, binarization at 170) - for faint/pattern PDFs."""
    if not HAS_OCR:
        return ""
    try:
        pages = convert_from_path(pdf_path, dpi=300, first_page=1, last_page=min(max_pages, 5))
        ocr_text = ""
        for page in pages:
            page = ImageOps.autocontrast(page.convert("L"))
            page = page.point(lambda x: 0 if x < 170 else 255, "1")  # Old script: bring out faint lines
            text = pytesseract.image_to_string(page, config="--psm 6 --oem 3", lang="eng")
            ocr_text += text + "\n"
        return ocr_text
    except Exception:
        return ""

def llm_vision_score_form(pdf_path):
    """Use vision LLM to score if PDF is the Section 301 form (for image-based PDFs)."""
    if not OLLAMA_VISION_MODEL:
        return None
    try:
        import base64
        from PIL import Image
        doc = fitz.open(pdf_path)
        page = doc[0]
        pix = page.get_pixmap(dpi=300, alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        if _is_mostly_blank_image(img):
            doc.close()
            return None  # Skip - vision hallucinates on blank/noise pages
        img = _preprocess_image_for_vision(img)
        doc.close()
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        b64 = base64.b64encode(buf.getvalue()).decode()
        r = requests.post(f"{OLLAMA_URL}/api/chat",
            json={"model": OLLAMA_VISION_MODEL, "messages": [{"role": "user", "content": VISION_IDENTIFY_PROMPT, "images": [b64]}], "stream": False},
            timeout=60)
        if r.status_code == 200:
            out = r.json().get("message", {}).get("content", "").strip()
            m = re.search(r'\b(\d{1,3})\b', out)
            if m:
                return min(100, max(0, int(m.group(1))))
    except Exception:
        pass
    return None

# Reuse extraction logic from section 301
def sanitize_for_excel(text):
    if not text:
        return ''
    text = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', str(text))
    return ''.join(c for c in text if ord(c) >= 32 or c in '\t\n\r').strip()

def normalize_bci_status(val):
    if not val:
        return ''
    v = str(val).strip().upper()
    if v == "P": return "Public Document"
    elif v == "PV": return "Public Version of BCI"
    elif v in ["B", "BCI"]: return "BCI"
    elif "PUBLIC VERSION" in v: return "Public Version of BCI"
    elif "PUBLIC" in v: return "Public Document"
    return val

def normalize_yes_no(val):
    if not val:
        return ''
    v = re.sub(r'\s+', '', str(val)).upper()
    if v in ["YES", "Y", "ON"]: return "YES"
    if v in ["NO", "N", "OFF"]: return "NO"
    if v in ["N/A", "NA"]: return "N/A"
    return val

# PDF font encoding fixes (scanned docs often have garbled text from CID/custom fonts)
# Order matters: longer patterns first
GARBLED_FIXES = [
    ('3FRVFTUPS*OGPSNBUJPO', 'Requestor Information'),
    ('3FRVFTUPS3FQSFTFOUBUJWF', 'Requestor Representative'),
    ('0SHBOJ[BUJPO/BNF', 'Organization Name'),
    ('3FRVFTUPS/BNF', 'Requestor Name'),
    ('0SHBOJ[BUJPO', 'Organization'),
    ('3FRVFTUPS', 'Requestor'),
    ('4FDUJPO*OWFTUJHBUJPO', 'Section Investigation'),
    ('4FDUJPO', 'Section'),
    ('*OWFTUJHBUJPO', 'Investigation'),
    ('/BNF', 'Name'),
    ('1MFBTFQSPWJEF', 'Please provide'),
    ('1MFBTF', 'Please'),
    ('QSPWJEF', 'provide'),
    ('JOEJDBUF', 'indicate'),
    ('1SPEVDU', 'Product'),
    ('&YDMVTJPO', 'Exclusion'),
    ('3FRVFTU', 'Request'),
    ('DPOUBJOT', 'contains'),
    ('CVTJOFTT', 'business'),
    ('DPOGJEFOUJBM', 'confidential'),
    ('EPDVNFOU', 'Document'),
    ('QVCMJD', 'Public'),
    ('WFSTJPO', 'version'),
    ('7&34*0/', 'VERSION'),
    ('$0/5*/6&%', 'CONTINUED'),
    ('#&-08', 'BELOW'),
    ('GPSN', 'form'),
    ('EJHJU', 'digit'),
    ('IJOT', 'HTSUS'),
    ('SFMBUJPOTIJQ', 'relationship'),
    ('BUUBDINFOUT', 'attachments'),
    ('EPNFTUJD', 'domestic'),
    ('BWBJMBCJMJUZ', 'availability'),
    ('HMPCBM', 'global'),
    ('WBMVF', 'value'),
    ('RVBOUJUZ', 'quantity'),
    ('DPNNFOUT', 'comments'),
    ('QVCMJDEPDVNFOU', 'Public Document'),
    ('QVCMJD7FSTJPO', 'Public Version'),
    ('7FSTJPOPG#$*', 'Version of BCI'),
    ('DPODFSO', 'concern'),
    ('QSPEVDU', 'product'),
    ('3FQSFTFOUBUJWF', 'Representative'),
    ('TVCNJUUJOH', 'submitting'),
    ('JOGPSNBUJPO', 'information'),
    ('*OWFTUJHBUJPO', 'Investigation'),
    ('$IJOB', 'China'),
    ('3FMBUFE', 'Related'),
    ('5FDIOPMPHZ', 'Technology'),
    ('5SBOTGFS', 'Transfer'),
    ('*OUFMMFDUVBM', 'Intellectual'),
    ('1SPQFSUZ', 'Property'),
    ('3FRVFTU', 'Request'),
    ('TQFDJGJFE', 'specified'),
    ('GJFMET', 'fields'),
    ('JNQPSUFS', 'Importer'),
]

def decode_garbled_pdf_text(text):
    """Fix garbled text from PDF font encoding (CID/custom fonts in scanned docs)."""
    if not text or len(text) < 50:
        return text
    for garbled, clean in GARBLED_FIXES:
        text = text.replace(garbled, clean)
    return text

def clean_garbled_text(text):
    if not text:
        return ''
    text = re.sub(r'^[I\|!]', '', text)
    text = re.sub(r'^[\.\,_\-\';:\|\s]+', '', text)
    text = re.sub(r'[_\-;:=]{2,}.*$', '', text)
    text = text.replace('_ln_c_', 'Inc').replace('ln_c', 'Inc')
    text = text.replace('IO_t_h_er', 'Other').replace('N-1-A', 'N/A')
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def is_scanned_pdf(pdf_path):
    """Detect image-based PDFs (scanned docs). Threshold 300 to catch hybrid/partial text."""
    try:
        doc = fitz.open(pdf_path)
        total_text = sum(len(p.get_text().strip()) for p in doc[:2])
        doc.close()
        return total_text < 300
    except:
        return False

def extract_from_widgets(doc, doc_id):
    result = {}
    has_widgets = False
    for page in doc:
        for w in page.widgets():
            has_widgets = True
            fn = (w.field_name or '').lower()
            fv = str(w.field_value or '').strip()
            if not fv or fv == "Off":
                continue
            if fn == 'bci':
                result["q1_bci_status"] = normalize_bci_status(fv)
            elif fn in ['prod name/descrip', 'product description']:
                result["q2_product_description"] = fv
            elif fn in ['hts 10 digit', 'hts', 'htsus']:
                code = re.sub(r"\D", "", fv)
                if 8 <= len(code) <= 10:
                    result["q3_htsus"] = code.ljust(10, '0')
            elif fn in ['petitioner name', 'requestor name']:
                result["q4_requestor_name"] = fv
            elif fn == 'organization name':
                result["q4_organization"] = fv
            elif fn in ['petitioner representative', 'requestor representative']:
                result["q4_representative"] = fv
            elif fn in ['relationship to prod', 'relationship']:
                result["q5_relationship"] = fv
            elif fn == 'attachments':
                result["q6_attachments"] = normalize_yes_no(fv)
            elif fn in ['bci attachment', 'attachment bci']:
                result["q7_attachment_bci"] = normalize_bci_status(fv)
            elif fn in ['dom avail', 'domestic avail']:
                result["q8_us_sources"] = normalize_yes_no(fv)
            elif fn in ['global avail', 'global availability']:
                result["q9_third_countries"] = normalize_yes_no(fv)
            elif fn in ['2015 value', '2015value']:
                result["q10_2015_value"] = fv.replace('$', '').replace(',', '')
            elif fn in ['2015 quant', '2015 quantity']:
                result["q10_2015_quantity"] = fv
            elif fn in ['2016 value', '2016value']:
                result["q10_2016_value"] = fv.replace('$', '').replace(',', '')
            elif fn in ['2016 quant', '2016 quantity']:
                result["q10_2016_quantity"] = fv
            elif fn in ['2017 value', '2017value']:
                result["q10_2017_value"] = fv.replace('$', '').replace(',', '')
            elif fn in ['2017 quant', '2017 quantity']:
                result["q10_2017_quantity"] = fv
            elif fn in ['comments', 'comment']:
                result["q11_supporting_info"] = fv
    return result, has_widgets

# --- Extraction pipeline: layout → extract text → clean garbled → match pattern words per question ---
# Pattern words that identify each question section (must be present to extract)
QUESTION_PATTERN_WORDS = {
    "q1_bci_status": ["public document", "public version", "bci", "business confidential", "1."],
    "q2_product_description": ["2.", "product", "concern", "please provide", "description"],
    "q3_htsus": ["3.", "10-digit", "10 digit", "htsus", "hts", "item number"],
    "q4_requestor_name": ["4.", "requestor", "name", "last", "first"],
    "q4_organization": ["organization", "name"],
    "q5_relationship": ["5.", "relationship", "product", "importer", "producer"],
    "q6_attachments": ["6.", "attachments"],
    "q7_attachment_bci": ["7.", "bci", "attachment"],
    "q8_us_sources": ["8.", "domestic", "avail", "us sources"],
    "q9_third_countries": ["9.", "global", "avail", "third countr"],
    "q10_2015_value": ["2015", "value"],
    "q10_2016_value": ["2016", "value"],
    "q10_2017_value": ["2017", "value"],
    "q11_supporting_info": ["11.", "supporting", "comments", "box"],
}


def _has_pattern_words(text, key, min_matches=1):
    """Check if pattern words for this question are present in text (validates form structure)."""
    if not text:
        return False
    t = text.lower()
    patterns = QUESTION_PATTERN_WORDS.get(key, [])
    matches = sum(1 for p in patterns if p in t)
    return matches >= min_matches


def _is_likely_form_text(text):
    """Quick check: does text have enough form structure to attempt extraction?"""
    if not text or len(text.strip()) < 80:
        return False
    t = text.lower()
    markers = ["section 301", "requestor", "product", "htsus", "10-digit", "10 digit", "relationship",
               "domestic", "2015", "2016", "2017", "exclusion", "form"]
    return sum(1 for m in markers if m in t) >= 1


def _is_likely_new_pattern_text(text):
    """Check for Primex/ECM-style letter format with product description, HTS subheading, Item:"""
    if not text or len(text.strip()) < 100:
        return False
    t = text.lower()
    has_hts = bool(re.search(r"hts\s+subheading\s*:\s*\d{4}\.\d{2}\.\d{4}", t))
    has_item = bool(re.search(r"item\s*:\s*\d+\w*\s+", t))
    has_product_section = "product description" in t or "i.\s*product" in t
    has_company = "primex" in t or "ecm industries" in t or "ecm " in t
    return (has_hts or has_item) and (has_company or has_product_section)


def extract_from_new_pattern_text(text_input, doc_id):
    """Extract from Primex/ECM-style letter format: I. Product Description, HTS subheading, Item:, etc."""
    result = {}
    text = _prepare_text_for_extraction(text_input)
    if not _is_likely_new_pattern_text(text):
        return result
    # Q3: HTS - "HTS subheading: 9025.19.8080" or "9025.19.8080"
    m = re.search(r"HTS\s+subheading\s*:\s*(\d{4})\.(\d{2})\.(\d{4})", text, re.I)
    if m:
        result["q3_htsus"] = m.group(1) + m.group(2) + m.group(3)
    if not result.get("q3_htsus"):
        m = re.search(r"\b(\d{4})\.(\d{2})\.(\d{4})\b", text)
        if m:
            result["q3_htsus"] = m.group(1) + m.group(2) + m.group(3)
    # Q2: Product - "Item: 00208CA 8" Thermometer Indoor and Outdoor" (stop before HTS subheading)
    m = re.search(r"Item\s*:\s*(\d+\w*\s+[^\n]+?)(?=\s*HTS\s+subheading|\n\n|\nItem\s+Description)", text, re.I)
    if m:
        result["q2_product_description"] = clean_garbled_text(m.group(1).strip())
    if not result.get("q2_product_description"):
        m = re.search(r"Item\s+Description\s*:\s*([^\n]{10,200})", text, re.I)
        if m:
            result["q2_product_description"] = clean_garbled_text(m.group(1).strip())
    if not result.get("q2_product_description"):
        m = re.search(r"We request that USTR exclude the ([^\.]+) that is imported", text, re.I)
        if m:
            result["q2_product_description"] = clean_garbled_text(m.group(1).strip())
    # Q4: Requestor - "Primex Family of Companies" or "ECM Industries"
    m = re.search(r"([A-Za-z\s]+(?:Family of Companies|Industries|Inc\.?|LLC|Corp))[^\n]*hereby submits", text)
    if m:
        org = clean_garbled_text(m.group(1).strip())
        if len(org) > 3:
            result["q4_organization"] = org
            result["q4_requestor_name"] = org.split("(")[0].strip() if "(" in org else org
    if not result.get("q4_organization"):
        if "primex" in text.lower():
            result["q4_organization"] = "Primex"
            result["q4_requestor_name"] = "Primex"
        elif "ecm industries" in text.lower() or re.search(r"ecm\s+industries", text, re.I):
            result["q4_organization"] = "ECM Industries"
            result["q4_requestor_name"] = "ECM Industries"
    # Q5: Relationship - often Importer in these letters
    if "importer" in text.lower() and text.lower().find("importer") < 3000:
        result["q5_relationship"] = "Importer"
    # Q8/Q9: "not available in the U.S. or countries outside of China"
    if re.search(r"not available in the U\.?S\.?|not available.*outside.*china", text, re.I):
        result["q8_us_sources"] = "NO"
        result["q9_third_countries"] = "NO"
    return result


def _extract_text_from_layout(doc_or_path, max_pages=5):
    """Layout analysis: get text from PDF blocks (preserves reading order)."""
    if isinstance(doc_or_path, str):
        doc = fitz.open(doc_or_path)
        close = True
    else:
        doc = doc_or_path
        close = False
    blocks = []
    try:
        for page in doc[:max_pages]:
            # get_text("dict") returns blocks with "lines" and "spans" - preserves layout
            d = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)
            for block in d.get("blocks", []):
                for line in block.get("lines", []):
                    for span in line.get("spans", []):
                        t = span.get("text", "").strip()
                        if t:
                            blocks.append(t)
    finally:
        if close:
            doc.close()
    return "\n".join(blocks)


def _prepare_text_for_extraction(text_input):
    """Step 1: Extract raw text. Step 2: Decode garbled. Step 3: Clean."""
    text = text_input if isinstance(text_input, str) else "\n".join([p.get_text() for p in text_input[:5]])
    text = decode_garbled_pdf_text(text)  # Fix PDF font encoding (scanned docs)
    text = re.sub(r'[~_=]+', ' ', text).replace('VERSION 1 CONTINUED BELOW', '')
    text = re.sub(r'\s+', ' ', text)
    return text


def extract_from_text(text_input, doc_id):
    """Pipeline: layout → extract text → clean garbled → match pattern words per question → extract value.
    Only extracts when pattern words are present (validates we're in the right section)."""
    result = {}
    text = _prepare_text_for_extraction(text_input)
    text_lower = text.lower()
    # Try standard form first
    if _is_likely_form_text(text):
        result = _extract_from_standard_form_text(text, text_lower, doc_id)
    # Fallback: new pattern (Primex/ECM letter format)
    elif _is_likely_new_pattern_text(text):
        result = extract_from_new_pattern_text(text_input, doc_id)
    return result


def _extract_from_standard_form_text(text, text_lower, doc_id):
    """Extract from standard Section 301 form structure."""
    result = {}
    # Q1: BCI status
    if _has_pattern_words(text, "q1_bci_status"):
        if "public document" in text_lower and "public version" not in text_lower[:800]:
            result["q1_bci_status"] = "Public Document"
        elif "public version" in text_lower or re.search(r'\bPV\b', text[:500]):
            result["q1_bci_status"] = "Public Version of BCI"
        elif re.search(r'1\..*?(P|PV|B|BCI)', text[:500], re.I):
            m = re.search(r'1\..*?(P|PV|B|BCI)', text[:500], re.I)
            result["q1_bci_status"] = normalize_bci_status(m.group(1))
    # Q2: Product description
    if _has_pattern_words(text, "q2_product_description"):
        m = re.search(r'2\.\s*(?:Please provide|provide).*?(?:concern|product)[:\s]*(.+?)(?=3\.|10-digit|10 digit)', text, re.I | re.S)
        if m:
            desc = clean_garbled_text(m.group(1).strip())
            if len(desc) > 15 and not re.search(r'^\d+\.', desc):
                result["q2_product_description"] = re.sub(r'\s+', ' ', desc)
    # Q3: HTSUS code
    if _has_pattern_words(text, "q3_htsus"):
        for match in re.finditer(r'\b(\d{10})\b', text):
            code = match.group(1)
            if code != '1023456789':
                result["q3_htsus"] = code
                break
        if not result.get("q3_htsus"):
            m = re.search(r'(\d{4})\.(\d{2})\.(\d{4})', text)
            if m:
                result["q3_htsus"] = m.group(1) + m.group(2) + m.group(3)
            else:
                m = re.search(r'(\d{4})[\.\s](\d{2})[\.\s](\d{2})[\.\s](\d{2,4})', text)
                if m:
                    result["q3_htsus"] = m.group(1) + m.group(2) + m.group(3) + m.group(4).zfill(2)
    # Q4: Requestor name & organization
    if _has_pattern_words(text, "q4_requestor_name"):
        m = re.search(r'4\.\s*Requestor Information.*?Name.*?:\s*([^\n]{3,100})', text, re.I | re.S)
        if m:
            name = clean_garbled_text(m.group(1))
            if len(name) > 2 and not re.search(r'organization|public|note', name, re.I):
                result["q4_requestor_name"] = name
        if not result.get("q4_requestor_name"):
            m = re.search(r'([A-Z][a-z]+,\s*[A-Z][a-z]{2,})\s*(?:\n|Organization|$)', text)
            if m:
                name = clean_garbled_text(m.group(1))
                if len(name) > 5 and 'organization' not in name.lower():
                    result["q4_requestor_name"] = name
        if not result.get("q4_requestor_name"):
            m = re.search(r'\d{4}\.\d{2}\.\d{4}\s*\n\s*([A-Z][a-z]+,\s*[A-Z][a-z]+)', text)
            if m:
                result["q4_requestor_name"] = clean_garbled_text(m.group(1))
    if _has_pattern_words(text, "q4_organization"):
        m = re.search(r'4\.\s*Requestor Information.*?Organization\s+Name[^\n:]*:\s*([^\n]+)', text, re.I | re.S)
        if m:
            result["q4_organization"] = clean_garbled_text(m.group(1))
        if not result.get("q4_organization"):
            m = re.search(r'[A-Z][a-z]+,\s*[A-Z][a-z]+\s*\n\s*([A-Za-z0-9\s,\.]+(?:Inc|LLC|Corp|Ltd)\.?)', text)
            if m:
                result["q4_organization"] = clean_garbled_text(m.group(1))
    # Q5: Relationship
    if _has_pattern_words(text, "q5_relationship"):
        m = re.search(r'5\..*?relationship.*?product[^\n]*:\s*([^\n]+)', text, re.I | re.S)
        if m:
            result["q5_relationship"] = clean_garbled_text(m.group(1))
        if not result.get("q5_relationship"):
            for rel in ["Importer", "U.S. Producer", "Other", "Manufacturer", "Distributor"]:
                if rel in text and text.find(rel) < 4000:
                    result["q5_relationship"] = rel
                    break
    # Q6-Q9
    for qnum, key in [("6", "q6_attachments"), ("7", "q7_attachment_bci"),
                      ("8", "q8_us_sources"), ("9", "q9_third_countries")]:
        if _has_pattern_words(text, key) and not result.get(key):
            m = re.search(qnum + r'\D{0,40}(YES|NO|N/?A|N\s*1\s*A)', text, re.I)
            if m:
                result[key] = normalize_yes_no(m.group(1))
    # Q10
    for year in ["2015", "2016", "2017"]:
        if _has_pattern_words(text, "q10_" + year + "_value"):
            m = re.search(year + r'\s*Value[:\s]*\$?\s*([\d,\.]+)', text, re.I)
            if m:
                result["q10_" + year + "_value"] = m.group(1).replace(',', '')
    # Q11
    if _has_pattern_words(text, "q11_supporting_info"):
        m = re.search(r'11\..*?box\).*?\n+(.{50,}?)(?=VERSION|END|$)', text, re.I | re.S)
        if m:
            result["q11_supporting_info"] = clean_garbled_text(re.sub(r'\s+', ' ', m.group(1).strip()))
    return result

def extract_all_fields(pdf, doc_id):
    fields = ["q1_bci_status", "q2_product_description", "q3_htsus", "q4_requestor_name", "q4_organization",
              "q4_representative", "q5_relationship", "q6_attachments", "q7_attachment_bci", "q8_us_sources",
              "q9_third_countries", "q10_2015_value", "q10_2015_quantity", "q10_2016_value", "q10_2016_quantity",
              "q10_2017_value", "q10_2017_quantity", "q11_supporting_info"]
    res = {f: "" for f in fields}
    res["doc_id"] = doc_id
    res["extraction_method"] = ""
    try:
        doc = fitz.open(pdf)
        w, has_widgets = extract_from_widgets(doc, doc_id)
        if has_widgets:
            for f in fields:
                res[f] = sanitize_for_excel(w.get(f, ""))
            res["extraction_method"] = "Form Fields"
        else:
            # Layout analysis: extract text from blocks (preserves reading order)
            raw_text = _extract_text_from_layout(doc, max_pages=5)
            if len(raw_text.strip()) < 100:
                raw_text = "\n".join([p.get_text() for p in doc[:5]])  # Fallback to plain get_text
            text = decode_garbled_pdf_text(raw_text)  # Fix PDF font encoding
            scanned = is_scanned_pdf(pdf)
            # Treat as scanned if little text OR explicitly detected (image PDFs)
            text_too_short = len(text.strip()) < 200
            # Garbled: font-encoded scanned PDFs (e.g. 2787) have raw text with patterns we decode
            text_has_garbled_patterns = any(g in raw_text for g, _ in GARBLED_FIXES[:20]) if raw_text else False
            text_looks_garbled = text_has_garbled_patterns and len(raw_text.strip()) > 300
            is_scanned = scanned or text_too_short or text_looks_garbled

            if is_scanned:
                # Scanned PDFs: try OCR extraction - many scanned forms are readable via OCR
                if HAS_OCR:
                    try:
                        ocr_text = _extract_ocr_text_for_scoring(pdf, max_pages=5)
                        ocr_text = decode_garbled_pdf_text(ocr_text)
                        if len(ocr_text.strip()) >= 200:
                            t = extract_from_text(ocr_text, doc_id)
                            for f in fields:
                                res[f] = sanitize_for_excel(t.get(f, ""))
                            res["extraction_method"] = "Scanned+OCR"
                            res["filled"] = sum(1 for f in fields if res[f])
                            res["status"] = "Success" if res["filled"] >= 5 else "Low"
                            doc.close()
                            return res
                    except Exception:
                        pass
                # OCR failed or not available - mark as Scanned
                res["extraction_method"] = "Scanned"
                res["Notes"] = "Scanned"
            else:
                # Native text PDF: LLM in normal mode, rule-based in FAST_MODE
                if not FAST_MODE and USE_LLM and text.strip():
                    llm_data = llm_extract_fields(text)
                    for f in fields:
                        res[f] = sanitize_for_excel(llm_data.get(f, ""))
                    res["extraction_method"] = "Text+LLM"
                else:
                    t = extract_from_text(text, doc_id)
                    for f in fields:
                        res[f] = sanitize_for_excel(t.get(f, ""))
                    res["extraction_method"] = "Text"
        doc.close()
        res["filled"] = sum(1 for f in fields if res[f])
        res["status"] = "Success" if res["filled"] >= 5 else "Low"
    except Exception as e:
        res["status"] = "Error"
        res["filled"] = 0
        res["extraction_method"] = str(e)
    return res

def fetch_document_list():
    """Fetch documents from regulations.gov API for USTR-2018-0032 with full attributes (Complete-11k schema)."""
    doc_list = []
    page = 1
    headers = {"X-Api-Key": API_KEY}
    while True:
        url = f"https://api.regulations.gov/v4/documents?filter[docketId]={DOCKET_ID}&page[size]=100&page[number]={page}&sort=documentId"
        try:
            r = requests.get(url, headers=headers, timeout=30)
            if r.status_code != 200:
                print(f"API error {r.status_code}: {r.text[:200]}")
                break
            data = r.json()
            docs = data.get("data", [])
            if not docs:
                break
            for d in docs:
                attrs = d.get("attributes", {})
                doc_list.append({
                    "documentId": d["id"],
                    "Document URL": f"https://www.regulations.gov/document/{d['id']}",
                    "Posted Date": attrs.get("postedDate"),
                    "Last Modified Date": attrs.get("lastModifiedDate"),
                    "Withdrawn": 1 if attrs.get("withdrawn") else 0,
                    "title": attrs.get("title", ""),
                })
            total = data.get("meta", {}).get("totalElements", 0)
            print(f"Fetched page {page}: {len(docs)} docs (total so far: {len(doc_list)}/{total})")
            if len(doc_list) >= total or len(docs) < 100:
                break
            page += 1
            time.sleep(0.5)
        except Exception as e:
            print(f"Error fetching docs: {e}")
            break
    return doc_list


def get_document_attributes(doc_id):
    """Fetch document attributes (Posted Date, Last Modified Date, Withdrawn, comment) from API."""
    headers = {"X-Api-Key": API_KEY}
    try:
        r = requests.get(f"https://api.regulations.gov/v4/documents/{doc_id}", headers=headers, timeout=15)
        if r.status_code != 200:
            return {}
        d = r.json().get("data", {})
        attrs = d.get("attributes", {})
        comment = attrs.get("comment") or ""
        if comment and isinstance(comment, str):
            comment = re.sub(r"<br\s*/?>", "\n", comment, flags=re.I)
        return {
            "Document URL": f"https://www.regulations.gov/document/{doc_id}",
            "Posted Date": attrs.get("postedDate"),
            "Last Modified Date": attrs.get("lastModifiedDate"),
            "Withdrawn": 1 if attrs.get("withdrawn") else 0,
            "comment": comment,
        }
    except Exception:
        return {"Document URL": f"https://www.regulations.gov/document/{doc_id}", "Posted Date": "", "Last Modified Date": "", "Withdrawn": None, "comment": ""}


def get_ustr_response_date(doc_id):
    """Extract USTR Response Date from Re_ attachment (add_date_col.py logic)."""
    if pdfplumber is None:
        return None
    headers = {"X-Api-Key": API_KEY}
    try:
        r = requests.get(f"https://api.regulations.gov/v4/documents/{doc_id}/attachments", headers=headers, timeout=15)
        if r.status_code != 200:
            return None
        for att in r.json().get("data", []):
            title = att.get("attributes", {}).get("title", "")
            if not (title.startswith("Re_") or "Re_" in title):
                continue
            file_formats = att.get("attributes", {}).get("fileFormats") or []
            if not file_formats:
                continue
            pdf_url = file_formats[0].get("fileUrl")
            if not pdf_url:
                continue
            # Download and extract date
            resp = requests.get(pdf_url, timeout=30)
            if resp.status_code != 200:
                continue
            try:
                with pdfplumber.open(io.BytesIO(resp.content)) as pdf:
                    text = (pdf.pages[0].extract_text() or "")[:1000]
                m = re.search(r"(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{1,2}),?\s+(\d{4})", text)
                if m:
                    date_str = m.group(0).replace(",", "")
                    parsed = datetime.strptime(date_str, "%B %d %Y")
                    return parsed.strftime("%Y-%m-%d")
            except Exception:
                pass
            break
    except Exception:
        pass
    return None

# Content patterns that identify the standard Section 301 exclusion request form
STANDARD_FORM_MARKERS = [
    r"section\s+301\s+investigation",  # Form title - strong signal
    r"form\s+to\s+request\s+exclusion",  # Form title
    r"requestor\s+information",
    r"10-?digit\s+hts",
    r"htsus?\s*code",
    r"relationship\s+to\s+(?:the\s+)?product",
    r"product\s+description",
    r"domestic\s+avail",
    r"global\s+(?:avail|availability)",
    r"2015\s*value",
    r"2016\s*value",
    r"2017\s*value",
    r"bci\s*(?:status)?",
    r"petitioner\s+name",
    r"organization\s+name",
]
STANDARD_FORM_FIELD_NAMES = ["bci", "hts", "htsus", "prod name", "product description", "requestor name",
                             "petitioner name", "organization name", "relationship", "dom avail", "global avail",
                             "2015 value", "2016 value", "2017 value"]

# Cover letter patterns: formal letter format, NOT the standard structured form
COVER_LETTER_INDICATORS = [
    r"Dear\s+(?:Ambassador|Sir|Madam|Mr\.|Ms\.|U\.S\.\s*Trade\s*Representative)",
    r"hereby\s+submits?\s+(?:this\s+)?(?:request|the\s+following)",
    r"hereby\s+submit\s+(?:the\s+following\s+)?supplemental",
    r"respectfully\s+(?:submits?|requests?)",
    r"Submitted\s+via\s+regulations\.gov",
    r"Office\s+of\s+(?:the\s+)?(?:US|U\.S\.)\s*Trade\s*Representative",
    r"600\s+17th\s+Street",
    r"Subject\s*:\s*Procedures\s+To\s+Consider",
    r"Re\s*:\s*Procedures\s+To\s+Consider",
    r"Re\s*:\s*Exclusion\s+(?:from\s+)?(?:additional\s+)?[Tt]ariff",  # DC Safety style
    r"Request\s+for\s+Exclusion\s+from\s+Section\s+301",
    r"Reference\s*:\s*Request\s+for\s+Exclusion",
    r"supplemental\s+comments?\s+related\s+to\s+tariffs",
    r"HTS\s+code\s+of\s+\d{4}\.\d{2}\.\d{4}",  # "HTS code of 8711.60.0000" (letter format)
]

# New-pattern: Primex/ECM filename "3917290090 ECM FLX-3410B P" or "9025198080 Primex 00315HDSB P"
NEW_PATTERN_FILENAME_RE = re.compile(r"^(\d{10})\s+(ECM|Primex)\s+([\w-]+)\s+P\s*$", re.I)
# Comment text indicating new-format submission
NEW_PATTERN_COMMENT_ECM = "ECM Industries LLC respectfully submits the attached exclusion request for HTS Code"
NEW_PATTERN_COMMENT_PRIMEX = "Primex Family of Companies respectfully submits the attached exclusion request for HTS Code"


def _is_new_pattern_by_metadata(url_titles, comment):
    """Check if document uses Primex/ECM new format from attachment titles or comment - no download needed."""
    comment = (comment or "").strip()
    for url, title in (url_titles or []):
        t = (title or "").strip()
        if NEW_PATTERN_FILENAME_RE.match(t):
            return True
    if NEW_PATTERN_COMMENT_ECM in comment or NEW_PATTERN_COMMENT_PRIMEX in comment:
        return True
    return False


def _extract_from_new_pattern_metadata(url_titles, comment):
    """Extract HTS, org, product from filename/comment - no download."""
    result = {}
    comment = (comment or "").strip()
    # From filename: "3917290090 ECM FLX-3410B P" -> HTS=3917290090, product=FLX-3410B
    for url, title in (url_titles or []):
        t = (title or "").strip()
        m = NEW_PATTERN_FILENAME_RE.match(t)
        if m:
            result["q3_htsus"] = m.group(1)
            result["q2_product_description"] = m.group(3)  # product code
            result["_first_url"] = url
            if "ECM" in m.group(2).upper():
                result["q4_organization"] = "ECM Industries LLC"
                result["q4_requestor_name"] = "ECM Industries LLC"
            else:
                result["q4_organization"] = "Primex Family of Companies"
                result["q4_requestor_name"] = "Primex"
            break
    # From comment if no filename match: "HTS Code 3917290090"
    if not result.get("q3_htsus") and comment:
        m = re.search(r"HTS\s+Code\s+(\d{10})", comment, re.I)
        if m:
            result["q3_htsus"] = m.group(1)
        if NEW_PATTERN_COMMENT_ECM in comment:
            result["q4_organization"] = "ECM Industries LLC"
            result["q4_requestor_name"] = "ECM Industries LLC"
        elif NEW_PATTERN_COMMENT_PRIMEX in comment:
            result["q4_organization"] = "Primex Family of Companies"
            result["q4_requestor_name"] = "Primex"
    return result


# New-pattern documents: Primex, ECM Industries, etc. - letter format with I. Product Description, HTS subheading
NEW_PATTERN_INDICATORS = [
    r"I\.\s*\n?\s*Product\s+Description",  # "I. Product Description"
    r"II\.\s*\n?\s*(?:Primex|ECM)\s+[Cc]annot\s+[Ss]ource",
    r"HTS\s+subheading\s*:\s*\d{4}\.\d{2}\.\d{4}",
    r"Item\s*:\s*\d+\w*\s+[\w\s\"']+",  # "Item: 00208CA 8" Thermometer..."
]


def _is_new_pattern(pdf_path):
    """Detect Primex/ECM-style documents - letter format with product description, HTS, etc. Extract from these."""
    try:
        doc = fitz.open(pdf_path)
        text = "\n".join([p.get_text() for p in doc[:5]])
        doc.close()
        text = decode_garbled_pdf_text(text)
        t = text.lower()
        # Must have product/HTS structure typical of new pattern
        has_section = bool(re.search(r"I\.\s*\n?\s*product\s+description", t, re.I))
        has_hts = bool(re.search(r"hts\s+subheading\s*:\s*\d{4}\.\d{2}\.\d{4}", t, re.I))
        has_company = "primex" in t or "ecm industries" in t or "ecm " in t
        has_item = bool(re.search(r"item\s*:\s*\d+\w*\s+", t, re.I))
        # New pattern: letter-style with structured product info (not standard form)
        if (has_hts or has_item) and (has_company or has_section):
            return True
        markers = sum(1 for pat in NEW_PATTERN_INDICATORS if re.search(pat, text, re.I))
        return markers >= 2
    except Exception:
        return False


def _page_contains_primex_logo(pdf_path):
    """Detect Primex logo (blue stylized P) on first page using template matching. Requires opencv-python."""
    if not HAS_CV2 or not os.path.isfile(PRIMEX_LOGO_PATH):
        return False
    try:
        template = cv2.imread(PRIMEX_LOGO_PATH)
        if template is None:
            return False
        doc = fitz.open(pdf_path)
        page = doc[0]
        pix = page.get_pixmap(dpi=150, alpha=False)
        img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.height, pix.width, pix.n)
        doc.close()
        if pix.n == 4:
            img = cv2.cvtColor(img, cv2.COLOR_RGBA2RGB)
        elif pix.n != 3:
            return False
        gray_page = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY)
        gray_tpl = cv2.cvtColor(template, cv2.COLOR_BGR2GRAY)
        # Multi-scale: logo may be different size in doc
        for scale in [0.5, 0.75, 1.0, 1.25]:
            w, h = int(gray_tpl.shape[1] * scale), int(gray_tpl.shape[0] * scale)
            if w < 20 or h < 20:
                continue
            tpl = cv2.resize(gray_tpl, (w, h))
            res = cv2.matchTemplate(gray_page, tpl, cv2.TM_CCOEFF_NORMED)
            if res.max() > 0.65:
                return True
        return False
    except Exception:
        return False


# USTR response letters (from USTR to requester) - NOT the exclusion request form
USTR_RESPONSE_INDICATORS = [
    r"The\s+purpose\s+of\s+this\s+letter\s+is\s+to\s+inform\s+you",
    r"EXECUTIVE\s+OFFICE\s+OF\s+THE\s+PRESIDENT\s*\n\s*OFFICE\s+OF\s+THE\s+UNITED\s+STATES\s+TRADE\s+REPRESENTATIVE",
    r"RE:\s+Product\s+Exclusion\s+Request\s+Number\s*:",
]


def _is_ustr_response_letter(pdf_path):
    """Detect USTR response letters (Re_ docs) by content - downloaded via fallback URL with no title."""
    try:
        doc = fitz.open(pdf_path)
        text = "\n".join([p.get_text() for p in doc[:2]])
        doc.close()
        text = decode_garbled_pdf_text(text)
        t = text
        markers = sum(1 for pat in USTR_RESPONSE_INDICATORS if re.search(pat, t, re.I))
        return markers >= 1
    except Exception:
        return False


# Product attachment sheets: "Attachment A - Product Name" with mostly images, minimal text - NOT the form
PRODUCT_ATTACHMENT_INDICATORS = [
    r"^Attachment\s+[A-Z]\s*[-–]\s*",  # "Attachment A - Slush Drape", "Attachment A – Monopolar Cords"
    r"Attachment\s+[A-Z]\s*[-–]\s*[^\n]{2,80}$",  # Title line only, rest is images
]

def _is_cover_letter_not_form(pdf_path):
    """Detect cover letters / formal submissions - NOT the standard structured form. Treat as pdf not present."""
    try:
        # New-pattern docs (Primex, ECM) have extractable info - do NOT reject
        if _is_new_pattern(pdf_path):
            return False
        doc = fitz.open(pdf_path)
        page1_text = doc[0].get_text() if len(doc) > 0 else ""
        text = "\n".join([p.get_text() for p in doc[:3]])
        doc.close()
        text = decode_garbled_pdf_text(text)
        page1_text = decode_garbled_pdf_text(page1_text)
        t = text.lower()
        p1 = page1_text.lower()
        letter_count = sum(1 for pat in COVER_LETTER_INDICATORS if re.search(pat, t, re.I))
        form_count = sum(1 for pat in STANDARD_FORM_MARKERS if re.search(pat, t, re.I))
        # "Dear Ambassador" + no "Requestor Information" = cover letter
        # Standard form has "4. Requestor Information" section; letters don't
        if re.search(r"dear\s+ambassador", t) and "requestor information" not in t:
            return True
        # Strong form structure (4+ markers) = definitely the form, don't reject
        if form_count >= 4:
            return False
        # Cover letter: 2+ letter indicators, and few/no form structure markers
        if letter_count >= 2 and form_count <= 1:
            return True
        # Strong letter opening with no form structure
        if letter_count >= 1 and form_count == 0 and len(text.strip()) > 200:
            return True
        # "Dear Ambassador" + supplemental/request format = cover letter (e.g. Lyft e-bikes)
        if letter_count >= 1 and re.search(r"supplemental|supplement\s+to\s+#\d", t, re.I) and form_count <= 2:
            return True
        return False
    except Exception:
        return False


def _is_product_attachment_not_form(pdf_path):
    """Detect product attachment sheets (Attachment A/B with mostly images) - NOT the standard form.
    These have minimal text ('Public Document', 'Attachment A - Product Name') and large product photos."""
    try:
        doc = fitz.open(pdf_path)
        text = "\n".join([p.get_text() for p in doc[:3]])
        doc.close()
        text = decode_garbled_pdf_text(text)
        text_len = len(text.strip())
        # Product attachment: "Attachment A - X" or "Attachment B - Y" with very little text (< 600 chars)
        # Standard form has 2000+ chars of structured questions
        if text_len < 600 and re.search(r"Attachment\s+[A-Z]\s*[-–]", text, re.I):
            return True
        # "Public Document" + "Attachment A" as main content = product spec sheet
        if text_len < 400 and "public document" in text.lower() and "attachment" in text.lower():
            return True
        return False
    except Exception:
        return False

def _score_by_layout(doc, text=""):
    """Score by PDF structure. Standard form has FILLABLE form fields (widgets), not just rectangles.
    Image-heavy PDFs (product photos) have rectangles from image borders - do NOT score those high."""
    score = 0
    rect_count = 0
    field_names = set()
    try:
        for page in doc[:3]:
            for w in page.widgets():
                fn = (w.field_name or "").lower()
                if fn:
                    field_names.add(fn)
            # Count rectangles from vector drawings (form boxes)
            try:
                drawings = page.get_drawings()
                for d in drawings:
                    items = d.get("items", [])
                    for it in items:
                        if it[0] == "re":  # rectangle
                            rect_count += 1
            except Exception:
                pass
    except Exception:
        pass
    # Form fields: strong signal - standard form is a FILLABLE PDF. Product attachments have no form fields.
    for fn in STANDARD_FORM_FIELD_NAMES:
        if any(fn in f for f in field_names):
            score += 10
    if field_names:
        score += 15
    # Rectangle points: (a) when we have form fields, or (b) when substantial text proves it's not an image sheet
    # Filled/flattened forms often have no widgets but retain form structure (rectangles) + content
    text_len = len((text or "").strip())
    allow_rect_points = field_names or text_len >= 500  # Substantial text = not product image sheet
    if allow_rect_points:
        if rect_count >= 20:
            score += 50
        elif rect_count >= 10:
            score += 30
        elif rect_count >= 5:
            score += 15
    # No form fields + minimal text = likely image sheet (product photos), cap score
    # Exception: text_len >= 500 means we have substantial content - not an image sheet
    if not field_names and text_len < 500:
        score = min(score, 20)  # Image-heavy docs get at most 20
    return min(score, 100)


def score_pdf_as_standard_form(pdf_path):
    """Score 0-100. Uses vision (layout), layout structure, then text - avoids text-matching for scanned/garbled PDFs.
    Vision fallback: when layout+text score would be low, try vision LLM to catch faint/pattern PDFs."""
    # Cover letters (e.g. Lyft e-bikes supplemental) must not score high - reject early
    if _is_cover_letter_not_form(pdf_path):
        return 0
    # Product attachment sheets (Attachment A/B with mostly images) - NOT the form
    if _is_product_attachment_not_form(pdf_path):
        return 0
    text = ""
    layout_score = 0
    try:
        doc = fitz.open(pdf_path)
        for page in doc[:3]:
            text += page.get_text() + "\n"
            for w in page.widgets():
                fn = (w.field_name or "").lower()
                if fn:
                    pass  # used by _score_by_layout
        # 1. Vision LLM (skipped in FAST_MODE - saves ~60s per PDF)
        if not FAST_MODE and OLLAMA_VISION_MODEL:
            vis_score = llm_vision_score_form(pdf_path)
            if vis_score is not None:
                doc.close()
                return vis_score
        # 2. Layout-based: form fields + rectangle count (no rectangle points without form fields)
        layout_score = _score_by_layout(doc, text)
        doc.close()
        if layout_score >= 50:
            return layout_score
    except Exception:
        return 0
    # 3. Text-based LLM (skipped in FAST_MODE)
    text = decode_garbled_pdf_text(text)
    if not FAST_MODE and USE_LLM and text.strip():
        llm_score = llm_identify_standard_form(text)
        if llm_score is not None:
            return llm_score
    # 4. Rule-based: layout + text markers
    score = layout_score
    text_lower = text.lower()
    for pat in STANDARD_FORM_MARKERS:
        if re.search(pat, text_lower, re.I):
            score += 5
    # 5. OCR FALLBACK: scanned/faint PDFs - autocontrast + binarization at 170
    if score < FORM_SCORE_THRESHOLD and HAS_OCR and len(text.strip()) < 500:
        ocr_text = _extract_ocr_text_for_scoring(pdf_path)
        if ocr_text.strip():
            ocr_decoded = decode_garbled_pdf_text(ocr_text)
            ocr_lower = ocr_decoded.lower()
            ocr_score = layout_score
            for pat in STANDARD_FORM_MARKERS:
                if re.search(pat, ocr_lower, re.I):
                    ocr_score += 5
            if USE_LLM and ocr_decoded.strip():
                llm_score = llm_identify_standard_form(ocr_decoded[:3000])
                if llm_score is not None and llm_score > ocr_score:
                    ocr_score = llm_score
            if ocr_score >= FORM_SCORE_THRESHOLD:
                return min(ocr_score, 100)
    # 6. VISION FALLBACK: faint/pattern PDFs - vision LLM can see form structure
    if score < FORM_SCORE_THRESHOLD and OLLAMA_VISION_MODEL and len(text.strip()) < 500:
        vis_score = llm_vision_score_form(pdf_path)
        if vis_score is not None and vis_score >= FORM_SCORE_THRESHOLD:
            return vis_score
    return min(score, 100)

def get_all_pdf_urls(doc_id):
    """Get all PDF attachment URLs and titles. Returns list of (url, title) - title may be empty."""
    headers = {"X-Api-Key": API_KEY}
    result = []
    seen = set()
    api_titles = []  # Track titles from API to detect "only Re_" case
    try:
        r = requests.get(f"https://api.regulations.gov/v4/documents/{doc_id}/attachments",
                        headers=headers, timeout=15)
        if r.status_code == 200:
            for att in r.json().get("data", []):
                title = att.get("attributes", {}).get("title", "") or ""
                if title:
                    api_titles.append(title)
                for f in (att.get("attributes", {}).get("fileFormats") or []):
                    if f.get("format") == "pdf" and f.get("fileUrl"):
                        u = f["fileUrl"]
                        if u not in seen:
                            result.append((u, title))
                            seen.add(u)
                        break
    except Exception:
        pass
    # Fallback: only add when API returned no data, or API had non-Re_ attachments
    # If API says ALL attachments are Re_, skip fallback (would download Re_ doc and wrongly score as Success)
    only_re_from_api = api_titles and all(
        str(t or "").strip().lower().startswith("re_") for t in api_titles
    )
    if not only_re_from_api:
        for n in range(1, 10):
            u = f"https://downloads.regulations.gov/{doc_id}/attachment_{n}.pdf"
            if u not in seen:
                result.append((u, ""))
                seen.add(u)
    return result

def download_pdf_to_path(url, path, retry=2):
    for attempt in range(retry):
        try:
            r = requests.get(url, timeout=45, headers={
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"
            })
            if r.status_code == 200 and len(r.content) > 1000:
                ct = r.headers.get("Content-Type", "")
                if "pdf" in ct or not ct:
                    with open(path, "wb") as f:
                        f.write(r.content)
                    return path
        except Exception:
            time.sleep(2)
    return None

def find_and_extract_standard_form(doc_id, doc_info=None):
    """Download all PDFs, identify the standard form by content, extract from it.
    Returns Complete-11k schema: documentId, Document URL, Posted Date, Last Modified Date,
    attachment_url, USTR Response Date, Withdrawn, q1-q11, extraction_method, pipeline_script, etc."""
    # Get document attributes (from cache or API)
    if doc_info is None:
        doc_info = get_document_attributes(doc_id)
    # Get USTR Response Date from Re_ attachment (skip if SKIP_USTR_RESPONSE_DATE for speed)
    ustr_date = None if SKIP_USTR_RESPONSE_DATE else get_ustr_response_date(doc_id)
    if not FAST_MODE:
        time.sleep(0.1)  # Rate limit when doing heavy LLM calls

    url_titles = get_all_pdf_urls(doc_id)
    url_titles_all = url_titles  # Keep for metadata fallback (before Re_ filter)
    comment = doc_info.get("comment") or ""
    # Comment indicates form not attached - pdf not present
    if re.search(r"submit\s+(?:the\s+)?(?:confidential\s+)?(?:request\s+)?separately|too\s+large\s+to\s+attach", comment, re.I):
        return _make_result(doc_id, doc_info, ustr_date, status="pdf not present", attachment_url="", extraction_method="", filled=0, form_match_score=0)

    # NEW PATTERN (Primex/ECM): detect from filename or comment - NO DOWNLOAD
    if _is_new_pattern_by_metadata(url_titles, comment):
        meta = _extract_from_new_pattern_metadata(url_titles, comment)
        att_url = meta.pop("_first_url", "")
        if not att_url:
            non_re = [(u, t) for u, t in url_titles if not (str(t or "").strip().lower().startswith("re_"))]
            att_url = non_re[0][0] if non_re else (url_titles[0][0] if url_titles else "")
        return _make_result(doc_id, doc_info, ustr_date, status="new pattern", attachment_url=att_url,
                           extraction_method="Metadata", filled=sum(1 for k, v in meta.items() if v),
                           form_match_score=100, Notes="New pattern", **meta)

    # Skip Re_ attachments (USTR response letters) - do not download
    url_titles = [(u, t) for u, t in url_titles if not (str(t or "").strip().lower().startswith("re_"))]
    # No non-Re_ attachments (only Re_ docs or no PDFs)
    if not url_titles:
        return _make_result(doc_id, doc_info, ustr_date, status="pdf not present", attachment_url="", extraction_method="", filled=0, form_match_score=0)

    # Score all PDFs; new-pattern docs (Primex/ECM) accepted and marked separately
    candidates = []
    for i, (url, title) in enumerate(url_titles):
        path = os.path.join(OUTPUT_DIR, f"temp_{doc_id.replace('/', '_')}_{i}.pdf")
        if download_pdf_to_path(url, path):
            score = section301_detection.score_pdf_as_section301_form(path)  # GitHub detection logic
            candidates.append((score, path, url))
    candidates.sort(key=lambda x: -x[0])
    if not candidates:
        return _make_result(doc_id, doc_info, ustr_date, status="pdf not present", attachment_url="", extraction_method="", filled=0, form_match_score=0)

    best_score, best_pdf, best_url = candidates[0]
    # Standard form not present - try OCR rescue on each candidate (scanned forms with poor text extraction)
    if best_score < FORM_SCORE_THRESHOLD and HAS_OCR:
        for score, path, url in candidates:
            try:
                ocr_text = _extract_ocr_text_for_scoring(path)
                if len(ocr_text.strip()) >= 200:
                    ocr_decoded = decode_garbled_pdf_text(ocr_text)
                    ocr_lower = ocr_decoded.lower()
                    ocr_markers = sum(1 for pat in STANDARD_FORM_MARKERS if re.search(pat, ocr_lower, re.I))
                    ocr_score = ocr_markers * 5
                    # 5+ markers = strong form signal (avoids false positives from cover letters)
                    threshold = 25 if ocr_markers >= 5 else FORM_SCORE_THRESHOLD
                    if ocr_score >= threshold:
                        best_score = min(ocr_score, 100)
                        best_pdf, best_url = path, url
                        break
            except Exception:
                pass
    # Still below threshold - pdf not present
    if best_score < FORM_SCORE_THRESHOLD:
        for _, path, _ in candidates:
            try:
                os.remove(path)
            except Exception:
                pass
        return _make_result(doc_id, doc_info, ustr_date, status="pdf not present", attachment_url="", extraction_method="", filled=0, form_match_score=best_score)

    res = extract_all_fields(best_pdf, doc_id)
    # If poor extraction, try other PDFs (skipped in FAST_MODE)
    if not FAST_MODE and res.get("filled", 0) < 5 and len(candidates) > 1:
        for score, path, url in candidates[1:]:
            try:
                res2 = extract_all_fields(path, doc_id)
                if res2.get("filled", 0) > res.get("filled", 0):
                    res = res2
                    best_url = url
                    best_score = score
                    best_pdf = path
                    if res.get("filled", 0) >= 5:
                        break
            except Exception:
                pass
    # Cover letter / product attachment check: best PDF is wrong type, NOT the standard form
    def _is_wrong_pdf_type(path):
        if _is_new_pattern(path):
            return False  # New pattern (Primex/ECM) is acceptable - we extract from it
        if _is_ustr_response_letter(path):
            return True  # USTR response (Re_) - downloaded via fallback with no title
        return _is_cover_letter_not_form(path) or _is_product_attachment_not_form(path)

    if _is_wrong_pdf_type(best_pdf):
        # Try other attachments - form may be in attachment_2, attachment_3, etc.
        for score, path, url in candidates[1:]:
            if score >= FORM_SCORE_THRESHOLD and not _is_wrong_pdf_type(path):
                res = extract_all_fields(path, doc_id)
                best_pdf, best_url, best_score = path, url, score
                break
        else:
            # No non-cover-letter PDF found
            for _, path, _ in candidates:
                try:
                    os.remove(path)
                except Exception:
                    pass
            return _make_result(doc_id, doc_info, ustr_date, status="pdf not present", attachment_url="", extraction_method="", filled=0, form_match_score=best_score)
    res = {k: v for k, v in res.items() if k != "doc_id"}
    # Mark new-pattern docs (Primex, ECM) - pdf present but different format
    if _is_new_pattern(best_pdf):
        res["status"] = "new pattern"
        res["Notes"] = "New pattern"
    elif _is_new_pattern_by_metadata(url_titles_all, comment):
        # Fallback: metadata says new pattern (e.g. API returned empty attachment titles)
        meta = _extract_from_new_pattern_metadata(url_titles_all, comment)
        for k, v in meta.items():
            if k != "_first_url" and v and not res.get(k):
                res[k] = v
        res["status"] = "new pattern"
        res["Notes"] = "New pattern"
    for _, path, _ in candidates:
        try:
            os.remove(path)
        except Exception:
            pass
    return _make_result(doc_id, doc_info, ustr_date, attachment_url=best_url, form_match_score=best_score, **res)


def _make_result(doc_id, doc_info, ustr_date, status="", attachment_url="", extraction_method="", filled=0, form_match_score=0, **extracted):
    """Build result row matching Complete-11k-id-attributes schema + pipeline_script."""
    doc_info = doc_info or {}
    fields = ["q1_bci_status", "q2_product_description", "q3_htsus", "q4_requestor_name", "q4_organization",
              "q4_representative", "q5_relationship", "q6_attachments", "q7_attachment_bci", "q8_us_sources",
              "q9_third_countries", "q10_2015_value", "q10_2015_quantity", "q10_2016_value", "q10_2016_quantity",
              "q10_2017_value", "q10_2017_quantity", "q11_supporting_info"]
    row = {
        "documentId": doc_id,
        "Document URL": doc_info.get("Document URL", f"https://www.regulations.gov/document/{doc_id}"),
        "Posted Date": doc_info.get("Posted Date", ""),
        "Last Modified Date": doc_info.get("Last Modified Date", ""),
        "attachment_url": attachment_url,
        "USTR Response Date": ustr_date or "",
        "Withdrawn": doc_info.get("Withdrawn"),
    }
    for f in fields:
        row[f] = extracted.get(f, "")
    row["extraction_method"] = extracted.get("extraction_method", extraction_method)
    row["pipeline_script"] = "section301_ustr_list2"  # Identifies which script produced this row
    # Withdrawn=1 overrides status
    if doc_info.get("Withdrawn") == 1:
        row["status"] = "withdrawn"
    else:
        row["status"] = extracted.get("status", status)
    row["filled"] = extracted.get("filled", filled)
    row["form_match_score"] = form_match_score
    # Notes: "Scanned" only when Low AND we have a pdf; "PDF not present" when no form/no pdf
    status_val = extracted.get("status", status)
    has_attachment = bool(attachment_url and str(attachment_url).strip())
    if status_val == "pdf not present" or not has_attachment:
        row["Notes"] = "PDF not present"
    elif status_val == "Low":
        row["Notes"] = "Scanned"
    else:
        row["Notes"] = extracted.get("Notes", "")
    return row


def process_one(doc_id, doc_info=None):
    return find_and_extract_standard_form(doc_id, doc_info)


def apply_highlighting_and_summary(excel_file, results):
    """Add Summary sheet and highlight scanned rows (red)."""
    if not HAS_OPENPYXL:
        return
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ws.title = "Extracted Data"
        red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")  # Red for Low (scanned)
        gray_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")   # Gray for pdf not present
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow for withdrawn
        orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange for new pattern
        status_col = notes_col = withdrawn_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == "status":
                status_col = col_idx
            elif cell.value == "Notes":
                notes_col = col_idx
            elif cell.value == "Withdrawn":
                withdrawn_col = col_idx
        low_count = 0
        if status_col is None:
            wb.close()
            return
        for row_idx in range(2, ws.max_row + 1):
            is_withdrawn = withdrawn_col and ws.cell(row=row_idx, column=withdrawn_col).value == 1
            if is_withdrawn:
                # Withdrawn=1: yellow highlight, status=withdrawn
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
                if status_col:
                    ws.cell(row=row_idx, column=status_col).value = "withdrawn"
            else:
                status_val = str(ws.cell(row=row_idx, column=status_col).value or "")
                is_low = status_val == "Low"
                is_pdf_not_present = status_val == "pdf not present"
                if is_low:
                    low_count += 1
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = red_fill
                    if notes_col:
                        ws.cell(row=row_idx, column=notes_col).value = "Scanned"
                elif is_pdf_not_present:
                    low_count += 1
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = gray_fill  # Gray, not red
                    if notes_col:
                        ws.cell(row=row_idx, column=notes_col).value = "PDF not present"
                elif status_val == "new pattern":
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = orange_fill
                    if notes_col:
                        ws.cell(row=row_idx, column=notes_col).value = "New pattern"
        summary_sheet = wb.create_sheet("Summary", 0)
        summary_sheet["A1"] = "Metric"
        summary_sheet["B1"] = "Count"
        summary_sheet["A1"].font = Font(bold=True, size=12)
        summary_sheet["B1"].font = Font(bold=True, size=12)
        row = 2
        summary_sheet["A" + str(row)] = "Total Documents"
        summary_sheet["B" + str(row)] = len(results)
        row += 2
        low_status_count = sum(1 for r in results if r.get("status") == "Low")
        pdf_not_present_count = sum(1 for r in results if r.get("status") == "pdf not present")
        summary_sheet["A" + str(row)] = "Scanned / PDF not present"
        summary_sheet["A" + str(row)].font = Font(bold=True)
        row += 1
        summary_sheet["A" + str(row)] = "Scanned (red highlight)"
        summary_sheet["B" + str(row)] = low_status_count
        summary_sheet.cell(row=row, column=1).fill = red_fill
        summary_sheet.cell(row=row, column=2).fill = red_fill
        row += 1
        summary_sheet["A" + str(row)] = "PDF not present (gray highlight)"
        summary_sheet["B" + str(row)] = pdf_not_present_count
        summary_sheet.cell(row=row, column=1).fill = gray_fill
        summary_sheet.cell(row=row, column=2).fill = gray_fill
        row += 1
        new_pattern_count = sum(1 for r in results if r.get("status") == "new pattern")
        summary_sheet["A" + str(row)] = "New pattern (orange highlight)"
        summary_sheet["B" + str(row)] = new_pattern_count
        summary_sheet.cell(row=row, column=1).fill = orange_fill
        summary_sheet.cell(row=row, column=2).fill = orange_fill
        row += 2
        withdrawn_count = sum(1 for r in results if r.get("Withdrawn") == 1)
        summary_sheet["A" + str(row)] = "Withdrawn"
        summary_sheet["A" + str(row)].font = Font(bold=True)
        row += 1
        summary_sheet["A" + str(row)] = "Withdrawn (yellow highlight)"
        summary_sheet["B" + str(row)] = withdrawn_count
        summary_sheet.cell(row=row, column=1).fill = yellow_fill
        summary_sheet.cell(row=row, column=2).fill = yellow_fill
        row += 2
        summary_sheet["A" + str(row)] = "Status"
        summary_sheet["A" + str(row)].font = Font(bold=True)
        row += 1
        success = sum(1 for r in results if r.get("status") == "Success")
        low = sum(1 for r in results if r.get("status") == "Low")
        withdrawn_status = sum(1 for r in results if r.get("status") == "withdrawn")
        summary_sheet["A" + str(row)] = "Success"
        summary_sheet["B" + str(row)] = success
        row += 1
        summary_sheet["A" + str(row)] = "Low"
        summary_sheet["B" + str(row)] = low
        row += 1
        summary_sheet["A" + str(row)] = "New pattern"
        summary_sheet["B" + str(row)] = new_pattern_count
        row += 1
        summary_sheet["A" + str(row)] = "Withdrawn"
        summary_sheet["B" + str(row)] = withdrawn_status
        row += 2
        form_count = sum(1 for r in results if "Form Fields" in str(r.get("extraction_method", "")))
        text_count = sum(1 for r in results if "Text" in str(r.get("extraction_method", "")) and "Scanned" not in str(r.get("extraction_method", "")))
        summary_sheet["A" + str(row)] = "Extraction Methods"
        summary_sheet["A" + str(row)].font = Font(bold=True)
        row += 1
        summary_sheet["A" + str(row)] = "Form Fields"
        summary_sheet["B" + str(row)] = form_count
        row += 1
        summary_sheet["A" + str(row)] = "Text / Text+LLM"
        summary_sheet["B" + str(row)] = text_count
        summary_sheet.column_dimensions["A"].width = 35
        summary_sheet.column_dimensions["B"].width = 15
        wb.save(excel_file)
        wb.close()
    except Exception as e:
        print(f"Warning: Could not add summary sheet: {e}")


def main():
    print("=" * 60)
    print("Section 301 List 2 (USTR-2018-0032) - Complete-11k schema output")
    print("=" * 60)
    doc_list = []
    form_fields_ids = set()
    existing_rows = {}

    # When EXISTING_EXCEL_PATH is set: use it as source of ALL doc IDs (skip Form Fields, process rest)
    if EXISTING_EXCEL_PATH and os.path.isfile(EXISTING_EXCEL_PATH):
        try:
            df_ex = pd.read_excel(EXISTING_EXCEL_PATH, sheet_name="Extracted Data")
        except Exception:
            df_ex = pd.read_excel(EXISTING_EXCEL_PATH)
        if "documentId" in df_ex.columns:
            seen = set()
            doc_list = []
            for _, row in df_ex.iterrows():
                did = str(row.get("documentId", ""))
                if did and did not in seen:
                    seen.add(did)
                    doc_list.append({"documentId": did})
            if "extraction_method" in df_ex.columns:
                mask = df_ex["extraction_method"].astype(str).str.contains("Form Fields", na=False)
                form_fields_ids = set(df_ex.loc[mask, "documentId"].astype(str).tolist())
            for _, row in df_ex.iterrows():
                existing_rows[str(row.get("documentId", ""))] = row.to_dict()
            print(f"Loaded existing: {EXISTING_EXCEL_PATH}")
            print(f"Total: {len(doc_list)} docs, Form Fields (skip): {len(form_fields_ids)}")
        else:
            print(f"Warning: existing Excel missing documentId column")
    if not doc_list:
        if os.path.exists("document_ids.txt"):
            with open("document_ids.txt") as f:
                ids = [l.strip() for l in f if l.strip() and not l.startswith("#")]
            doc_list = [{"documentId": did} for did in ids]
            print(f"Loaded {len(doc_list)} IDs from document_ids.txt")
        if not doc_list:
            print("Fetching document list from regulations.gov API...")
            doc_list = fetch_document_list()
    if not doc_list:
        print("No documents found.")
        return
    if END_INDEX is not None:
        doc_list = doc_list[START_INDEX:END_INDEX]
    elif START_INDEX > 0:
        doc_list = doc_list[START_INDEX:]
    if LIMIT:
        doc_list = doc_list[:LIMIT]

    doc_list_to_process = [d for d in doc_list if str(d["documentId"]) not in form_fields_ids]
    print(f"\nProcessing {len(doc_list_to_process)} documents (skipping {len(form_fields_ids)} Form Fields)...")
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(process_one, d["documentId"], d if "Posted Date" in d or "Document URL" in d else None): d["documentId"] for d in doc_list_to_process}
        for i, future in enumerate(as_completed(futures)):
            res = future.result()
            results.append(res)
            print(f"  {res.get('documentId')}: {res.get('status')} (score={res.get('form_match_score', 'N/A')})")
            if (i + 1) % 50 == 0:
                print(f"  Done {i + 1}/{len(doc_list_to_process)}")
    # Merge: Form Fields rows (from existing) + newly processed results
    merged = []
    results_by_id = {str(r.get("documentId")): r for r in results}
    for d in doc_list:
        did = str(d["documentId"])
        if did in form_fields_ids and did in existing_rows:
            merged.append(existing_rows[did])
        elif did in results_by_id:
            merged.append(results_by_id[did])
    results = merged
    df = pd.DataFrame(results)
    # Column order matching Complete-11k-id-attributes
    col_order = [
        "documentId", "Document URL", "Posted Date", "Last Modified Date",
        "attachment_url", "USTR Response Date", "Withdrawn",
        "q1_bci_status", "q2_product_description", "q3_htsus", "q4_requestor_name", "q4_organization",
        "q4_representative", "q5_relationship", "q6_attachments", "q7_attachment_bci",
        "q8_us_sources", "q9_third_countries",
        "q10_2015_value", "q10_2015_quantity", "q10_2016_value", "q10_2016_quantity",
        "q10_2017_value", "q10_2017_quantity", "q11_supporting_info",
        "extraction_method", "pipeline_script", "status", "filled", "form_match_score", "Notes",
    ]
    df = df[[c for c in col_order if c in df.columns]]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = os.path.join(OUTPUT_DIR, f"section301_list2_{ts}.xlsx")
    df.to_excel(out_file, index=False)
    apply_highlighting_and_summary(out_file, results)
    print(f"\nOutput: {out_file}")
    print(f"Total: {len(results)} (Form Fields kept: {len(form_fields_ids)}, Re-processed: {len(doc_list_to_process)})")
    print(f"Success: {sum(1 for r in results if r.get('status') == 'Success')}")
    print("Columns: Complete-11k schema + pipeline_script=section301_ustr_list2")
    print("=" * 60)

if __name__ == "__main__":
    main()
