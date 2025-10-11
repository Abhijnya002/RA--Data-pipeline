#!/usr/bin/env python3
import pandas as pd
import requests
import fitz
import os
import re
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

try:
    import pytesseract
    from pdf2image import convert_from_path
    from PIL import ImageEnhance, ImageOps
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

START_ROW = 1
END_ROW = 1000
INPUT_FILE = "1-1000IDs_with_date_+_withdrawn_col.xlsx"
MAX_WORKERS = 5
DEBUG_DIR = "debug_extractions"
os.makedirs(DEBUG_DIR, exist_ok=True)

def download_pdf(url, doc_id, retry=2):
    for attempt in range(retry):
        try:
            r = requests.get(url, timeout=30)
            if r.status_code == 200:
                fname = "temp_" + doc_id.replace('/', '_') + ".pdf"
                with open(fname, "wb") as f:
                    f.write(r.content)
                return fname
        except:
            pass
    return None

def sanitize_for_excel(text):
    if not text:
        return ''
    text = re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', str(text))
    result = ''
    for char in text:
        if ord(char) >= 32 or char in '\t\n\r':
            result += char
    return result.strip()

def normalize_bci_status(val):
    if not val:
        return ''
    v = str(val).strip().upper()
    if v == "P":
        return "Public Document"
    elif v == "PV":
        return "Public Version of BCI"
    elif v == "B" or v == "BCI":
        return "BCI"
    elif "PUBLIC VERSION" in v:
        return "Public Version of BCI"
    elif "PUBLIC" in v:
        return "Public Document"
    return val

def normalize_yes_no(val):
    if not val:
        return ''
    v = re.sub(r'\s+', '', str(val)).upper()
    if v in ["YES","Y","ON"]:
        return "YES"
    if v in ["NO","N","OFF"]:
        return "NO"
    if v in ["N/A","NA"]:
        return "N/A"
    return val

def clean_garbled_text(text):
    if not text:
        return ''
    text = re.sub(r'^[I\|!]', '', text)
    text = re.sub(r'^[\.\,_\-\';:\|\s]+', '', text)
    text = re.sub(r'[_\-;:=]{2,}.*$', '', text)
    text = re.sub(r'[~=]{3,}', '', text)
    text = re.sub(r"['\-\.;:]+", '', text)
    text = text.replace('_ln_c_', 'Inc')
    text = text.replace('ln_c', 'Inc')
    text = text.replace('*ln*c', 'Inc')
    text = text.replace('IO_t_h_er', 'Other')
    text = text.replace('N-1-A', 'N/A')
    text = text.replace('N1A', 'N/A')
    text = re.sub(r'\.{2,}', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

def is_scanned_pdf(pdf_path):
    try:
        doc = fitz.open(pdf_path)
        total_text = 0
        for page in doc[:2]:
            total_text += len(page.get_text().strip())
        doc.close()
        return total_text < 200
    except:
        return False

def extract_ocr_text(pdf_path):
    if not HAS_OCR:
        return ""
    try:
        pages = convert_from_path(pdf_path, dpi=300, first_page=1, last_page=3)
        ocr_text = ""
        for page in pages:
            page = ImageOps.autocontrast(page.convert('L'))
            page = page.point(lambda x: 0 if x < 170 else 255, '1')
            text = pytesseract.image_to_string(page, config='--psm 6 --oem 3', lang='eng')
            ocr_text += text + "\n"
        return ocr_text
    except Exception as e:
        print("  OCR error: " + str(e))
        return ""

def extract_from_widgets(doc, doc_id):
    result = {}
    has_widgets = False
    
    for page in doc:
        for w in page.widgets():
            has_widgets = True
            fn = (w.field_name or '').lower()
            fv = str(w.field_value or '').strip()
            
            if not fv or fv == "Off":
                continue
            
            if fn == 'bci':
                result["q1_bci_status"] = normalize_bci_status(fv)
            elif fn in ['prod name/descrip', 'product description']:
                result["q2_product_description"] = fv
            elif fn in ['hts 10 digit', 'hts', 'htsus']:
                code = re.sub(r"\D", "", fv)
                if 8 <= len(code) <= 10:
                    result["q3_htsus"] = code.ljust(10,'0')
            elif fn in ['petitioner name', 'requestor name']:
                result["q4_requestor_name"] = fv
            elif fn == 'organization name':
                result["q4_organization"] = fv
            elif fn in ['petitioner representative', 'requestor representative']:
                result["q4_representative"] = fv
            elif fn in ['relationship to prod', 'relationship']:
                result["q5_relationship"] = fv
            elif fn == 'attachments':
                result["q6_attachments"] = normalize_yes_no(fv)
            elif fn in ['bci attachment', 'attachment bci']:
                result["q7_attachment_bci"] = normalize_bci_status(fv)
            elif fn in ['dom avail', 'domestic avail']:
                result["q8_us_sources"] = normalize_yes_no(fv)
            elif fn in ['global avail', 'global availability']:
                result["q9_third_countries"] = normalize_yes_no(fv)
            elif fn in ['2015 value', '2015value']:
                result["q10_2015_value"] = fv.replace('$','').replace(',','')
            elif fn in ['2015 quant', '2015 quantity']:
                result["q10_2015_quantity"] = fv
            elif fn in ['2016 value', '2016value']:
                result["q10_2016_value"] = fv.replace('$','').replace(',','')
            elif fn in ['2016 quant', '2016 quantity']:
                result["q10_2016_quantity"] = fv
            elif fn in ['2017 value', '2017value']:
                result["q10_2017_value"] = fv.replace('$','').replace(',','')
            elif fn in ['2017 quant', '2017 quantity']:
                result["q10_2017_quantity"] = fv
            elif fn in ['comments', 'comment']:
                result["q11_supporting_info"] = fv
    
    return result, has_widgets

def extract_from_text(text_input, doc_id):
    result = {}
    
    if isinstance(text_input, str):
        text = text_input
    else:
        text = "\n".join([p.get_text() for p in text_input[:3]])
    
    original_text = text
    clean_text = re.sub(r'[~_=]+', ' ', text)
    clean_text = clean_text.replace('VERSION 1 CONTINUED BELOW', '')
    clean_text = re.sub(r'\s+', ' ', clean_text)
    
    debug_file = os.path.join(DEBUG_DIR, doc_id + ".txt")
    with open(debug_file, "w", encoding="utf-8") as f:
        f.write(original_text)
    
    # Q1 - More specific to avoid matching wrong sections
    if "Public Document" in text:
        result["q1_bci_status"] = "Public Document"
    elif re.search(r'1\..*?(P|PV|B|BCI)', text[:500], re.I):
        m = re.search(r'1\..*?(P|PV|B|BCI)', text[:500], re.I)
        result["q1_bci_status"] = normalize_bci_status(m.group(1))
    
    # Q2 - More specific pattern to avoid grabbing question labels
    m = re.search(r'2\.\s*Please provide.*?concern[:\s]*(.+?)(?=3\.|10-digit)', text, re.I|re.S)
    if m:
        desc = m.group(1).strip()
        if len(desc) > 50 and not re.search(r'^\d+\.', desc):
            result["q2_product_description"] = re.sub(r'\s+', ' ', desc)
    
    # Q3
    for match in re.finditer(r'\b(\d{10})\b', text):
        code = match.group(1)
        if code != '1023456789':
            result["q3_htsus"] = code
            break
    
    # Q4 - More specific to find actual name in Requestor Information section
    m = re.search(r'4\.\s*Requestor Information.*?Name.*?:\s*([^\n]{3,100})', text, re.I|re.S)
    if m:
        name = clean_garbled_text(m.group(1))
        if len(name) > 2 and not re.search(r'organization|public|note', name, re.I):
            result["q4_requestor_name"] = name
    
    # Q4 - Organization
    m = re.search(r'4\.\s*Requestor Information.*?Organization\s+Name[^\n:]*:\s*([^\n]+)', text, re.I|re.S)
    if m:
        org = clean_garbled_text(m.group(1))
        if len(org) > 2 and not re.search(r'note|representative', org, re.I):
            result["q4_organization"] = org
    
    # Q4 - Representative
    m = re.search(r'4\.\s*Requestor Information.*?Representative[^\n:]*:\s*([^\n]+)', text, re.I|re.S)
    if m:
        rep = clean_garbled_text(m.group(1))
        if rep and not re.match(r'N\s*/?\s*A', rep, re.I) and len(rep) > 1:
            result["q4_representative"] = rep
    
    # Q5
    m = re.search(r'5\..*?relationship.*?product[^\n]*:\s*([^\n]+)', text, re.I|re.S)
    if m:
        rel = clean_garbled_text(m.group(1))
        if rel and len(rel) > 1 and not re.search(r'CONTINUED|^\d+\.', rel):
            result["q5_relationship"] = rel
    
    # Q6-Q9 with expanded windows
    if "6." in clean_text and "7." in clean_text:
        section = clean_text[clean_text.find("6."):clean_text.find("7.")]
    elif "6." in clean_text:
        section = clean_text[clean_text.find("6."):clean_text.find("6.") + 600]
    else:
        section = ""
    
    if section:
        if "NO" in section and "YES" not in section:
            result["q6_attachments"] = "NO"
        elif "YES" in section:
            result["q6_attachments"] = "YES"
        elif "N/A" in section:
            result["q6_attachments"] = "N/A"
    
    if "7." in clean_text and "8." in clean_text:
        section = clean_text[clean_text.find("7."):clean_text.find("8.")]
    elif "7." in clean_text:
        section = clean_text[clean_text.find("7."):clean_text.find("7.") + 600]
    else:
        section = ""
    
    if section:
        if "NO" in section and "YES" not in section:
            result["q7_attachment_bci"] = "NO"
        elif "YES" in section:
            result["q7_attachment_bci"] = "YES"
        elif "N/A" in section or "NJA" in section or "jNJA" in section:
            result["q7_attachment_bci"] = "N/A"
        else:
            m = re.search(r'\b(P|PV|B)\b', section)
            if m:
                result["q7_attachment_bci"] = normalize_bci_status(m.group(1))
    
    if "8." in clean_text and "9." in clean_text:
        section = clean_text[clean_text.find("8."):clean_text.find("9.")]
    elif "8." in clean_text:
        section = clean_text[clean_text.find("8."):clean_text.find("8.") + 600]
    else:
        section = ""
    
    if section:
        if "NO" in section and "YES" not in section:
            result["q8_us_sources"] = "NO"
        elif "YES" in section:
            result["q8_us_sources"] = "YES"
        elif "N/A" in section:
            result["q8_us_sources"] = "N/A"
    
    if "9." in clean_text and "10." in clean_text:
        section = clean_text[clean_text.find("9."):clean_text.find("10.")]
    elif "9." in clean_text:
        section = clean_text[clean_text.find("9."):clean_text.find("9.") + 600]
    else:
        section = ""
    
    if section:
        if "NO" in section and "YES" not in section:
            result["q9_third_countries"] = "NO"
        elif "YES" in section:
            result["q9_third_countries"] = "YES"
        elif "N/A" in section:
            result["q9_third_countries"] = "N/A"
    
    # Fallback for Q6-Q9
    for qnum, key in [("6", "q6_attachments"), ("7", "q7_attachment_bci"),
                      ("8", "q8_us_sources"), ("9", "q9_third_countries")]:
        if not result.get(key):
            pattern = qnum + r'\D{0,40}(YES|NO|N/?A|N\s*1\s*A|NJA|jNJA|P|PV|B)'
            m = re.search(pattern, text, re.I)
            if m:
                v = m.group(1).upper().replace(" ", "")
                if "1A" in v or "JA" in v:
                    v = "N/A"
                elif v.startswith("Y"):
                    v = "YES"
                elif v in ["N", "NO"]:
                    v = "NO"
                elif v in ["P", "PV", "B"]:
                    v = normalize_bci_status(v)
                result[key] = v
    
    # Q10
    for year in ["2015", "2016", "2017"]:
        pattern = year + r'\s*Value[:\s]*\$?\s*([\d,\.]+)\s*(Million|Thousand)?'
        m = re.search(pattern, text, re.I)
        if m:
            val = m.group(1).replace(',', '')
            if m.group(2) and 'million' in m.group(2).lower():
                try:
                    val = str(int(float(m.group(1).replace(',', '')) * 1000000))
                except:
                    pass
            result["q10_" + year + "_value"] = val
    
    for year in ["2015", "2016", "2017"]:
        pattern = year + r'\s*Quant[^\n]*\n+([^\n]+)'
        m = re.search(pattern, text, re.I)
        if m:
            result["q10_" + year + "_quantity"] = m.group(1).strip()
    
    # Q11 - No character limit
    m = re.search(r'11\..*?box\).*?\n+(.{100,}?)(?=VERSION|END)', text, re.I|re.S)
    if m:
        info = re.sub(r'\s+', ' ', m.group(1).strip())
        if not re.search(r'please provide information', info[:50], re.I):
            result["q11_supporting_info"] = info
    
    return result

def extract_all_fields(pdf, doc_id):
    fields = [
        "q1_bci_status", "q2_product_description", "q3_htsus",
        "q4_requestor_name", "q4_organization", "q4_representative",
        "q5_relationship", "q6_attachments", "q7_attachment_bci",
        "q8_us_sources", "q9_third_countries",
        "q10_2015_value", "q10_2015_quantity",
        "q10_2016_value", "q10_2016_quantity",
        "q10_2017_value", "q10_2017_quantity",
        "q11_supporting_info"
    ]
    
    res = {}
    for f in fields:
        res[f] = ""
    res["doc_id"] = doc_id
    res["extraction_method"] = ""
    
    try:
        doc = fitz.open(pdf)
        w, has_widgets = extract_from_widgets(doc, doc_id)
        
        if has_widgets:
            print("  Form")
            for f in fields:
                res[f] = sanitize_for_excel(w.get(f, ""))
            res["extraction_method"] = "Form Fields"
        else:
            scanned = is_scanned_pdf(pdf)
            
            if scanned and HAS_OCR:
                print("  OCR")
                ocr_text = extract_ocr_text(pdf)
                t = extract_from_text(ocr_text, doc_id)
            else:
                print("  Text")
                t = extract_from_text(doc, doc_id)
            
            res["extraction_method"] = "Scanned (OCR)"
            
            for f in fields:
                res[f] = sanitize_for_excel(t.get(f, ""))
        
        doc.close()
        
        filled = sum(1 for f in fields if res[f])
        res["filled"] = filled
        res["status"] = "Success" if filled >= 5 else "Low"
        
    except Exception as e:
        res["status"] = "Error"
        res["filled"] = 0
        res["extraction_method"] = "Error"
    
    return res

def apply_highlighting_and_summary(excel_file, results):
    wb = load_workbook(excel_file)
    ws = wb.active
    ws.title = "Extracted Data"
    
    light_red = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    dark_red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    method_col = None
    withdrawn_col = None
    
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == "extraction_method":
            method_col = col_idx
        elif cell.value == "Withdrawn":
            withdrawn_col = col_idx
    
    scanned_count = 0
    if method_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=method_col)
            if cell.value and "Scanned" in str(cell.value):
                scanned_count += 1
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = light_red
    
    withdrawn_count = 0
    if withdrawn_col:
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=withdrawn_col)
            if cell.value == 1:
                withdrawn_count += 1
                cell.fill = dark_red
    
    summary_sheet = wb.create_sheet("Summary", 0)
    
    form_count = len([r for r in results if r.get('extraction_method') == 'Form Fields'])
    scanned_ocr_count = len([r for r in results if r.get('extraction_method') == 'Scanned (OCR)'])
    success_count = len([r for r in results if r.get('status') == 'Success'])
    low_count = len([r for r in results if r.get('status') == 'Low'])
    error_count = len([r for r in results if r.get('status') == 'Error'])
    no_url_count = len([r for r in results if r.get('attachment_url') == 'nan'])
    avg_filled = sum(r.get('filled', 0) for r in results) / len(results) if results else 0
    
    summary_sheet['A1'] = 'Metric'
    summary_sheet['B1'] = 'Count'
    summary_sheet['A1'].font = Font(bold=True, size=12)
    summary_sheet['B1'].font = Font(bold=True, size=12)
    
    row = 2
    summary_sheet['A' + str(row)] = 'Total Documents'
    summary_sheet['B' + str(row)] = len(results)
    
    row += 2
    summary_sheet['A' + str(row)] = 'Extraction Methods'
    summary_sheet['A' + str(row)].font = Font(bold=True)
    
    row += 1
    summary_sheet['A' + str(row)] = 'Form Fields'
    summary_sheet['B' + str(row)] = form_count
    
    row += 1
    summary_sheet['A' + str(row)] = 'Scanned (OCR)'
    summary_sheet['B' + str(row)] = scanned_ocr_count
    summary_sheet.cell(row=row, column=1).fill = light_red
    summary_sheet.cell(row=row, column=2).fill = light_red
    
    row += 2
    summary_sheet['A' + str(row)] = 'Status Breakdown'
    summary_sheet['A' + str(row)].font = Font(bold=True)
    
    row += 1
    summary_sheet['A' + str(row)] = 'Success'
    summary_sheet['B' + str(row)] = success_count
    
    row += 1
    summary_sheet['A' + str(row)] = 'Low Extraction'
    summary_sheet['B' + str(row)] = low_count
    
    row += 1
    summary_sheet['A' + str(row)] = 'Error'
    summary_sheet['B' + str(row)] = error_count
    
    row += 1
    summary_sheet['A' + str(row)] = 'No URL'
    summary_sheet['B' + str(row)] = no_url_count
    
    row += 2
    summary_sheet['A' + str(row)] = 'Withdrawn Documents'
    summary_sheet['B' + str(row)] = withdrawn_count
    summary_sheet.cell(row=row, column=2).fill = dark_red
    
    row += 2
    summary_sheet['A' + str(row)] = 'Average Fields Filled'
    summary_sheet['B' + str(row)] = str(round(avg_filled, 1)) + '/18'
    
    summary_sheet.column_dimensions['A'].width = 30
    summary_sheet.column_dimensions['B'].width = 15
    
    wb.save(excel_file)
    wb.close()
    return scanned_count, withdrawn_count

def process_single(rownum, doc_id, url, source_row):
    out = {"row_number": rownum, "doc_id": doc_id}
    
    out["Document_URL"] = source_row.get("Document URL", "")
    out["USTR_Response_Date"] = source_row.get("USTR Response Date", "")
    out["Withdrawn"] = source_row.get("Withdrawn", "")
    
    if not url or pd.isna(url) or str(url).strip() == '':
        out["attachment_url"] = "no_url"
        out["status"] = "No URL"
        out["extraction_method"] = ""
        out["filled"] = 0
        return out
    
    out["attachment_url"] = url
    pdf = download_pdf(url, doc_id)
    if not pdf:
        out["status"] = "Failed"
        out["extraction_method"] = ""
        out["filled"] = 0
        return out
    
    out.update(extract_all_fields(pdf, doc_id))
    try:
        os.remove(pdf)
    except:
        pass
    return out

def main():
    df = pd.read_excel(INPUT_FILE)
    batch = df.iloc[START_ROW-1:END_ROW]
    tasks = []
    for i, r in batch.iterrows():
        doc_id = str(r.get("documentId", ""))
        url = str(r.get("attachment_url", ""))
        tasks.append((i+1, doc_id, url, r.to_dict()))
    
    results = []
    print("Processing " + str(len(tasks)) + " documents...")
    
    executor = ThreadPoolExecutor(MAX_WORKERS)
    futures = {}
    for t in tasks:
        future = executor.submit(process_single, *t)
        futures[future] = t
    
    for future in as_completed(futures):
        results.append(future.result())
    
    executor.shutdown()
    
    results.sort(key=lambda x: x["row_number"])
    df_out = pd.DataFrame(results)
    
    cols = [c for c in df_out.columns if c not in ['extraction_method', 'Document_URL', 'USTR_Response_Date', 'Withdrawn']]
    cols.extend(['Document_URL', 'USTR_Response_Date', 'Withdrawn', 'extraction_method'])
    df_out = df_out[[c for c in cols if c in df_out.columns]]
    
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = "section301_" + str(START_ROW) + "_" + str(END_ROW) + "_" + ts + ".xlsx"
    df_out.to_excel(out, index=False)
    
    scanned_count, withdrawn_count = apply_highlighting_and_summary(out, results)
    
    success = len([r for r in results if r.get('status') == 'Success'])
    
    print("\n" + "="*60)
    print("Output: " + out)
    print("Total: " + str(len(results)))
    print("Success: " + str(success))
    print("Scanned (OCR): " + str(scanned_count) + " <- Light red")
    print("Withdrawn: " + str(withdrawn_count) + " <- Dark red")
    print("\nSee 'Summary' sheet for statistics")
    print("="*60)

if __name__ == "__main__":
    main()