"""
Section 301 data extraction - uses section301_detection (GitHub data-extraction.py logic).
Reads IDs from txt file, parallel workers, NO LLM. Adds Summary sheet.
"""
import pandas as pd
import requests
import os
import time
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import section301_detection  # GitHub data-extraction.py - main and only PDF form detection

# Rate limit tracking (shared across workers)
_rate_limit_count = 0
_rate_limit_lock = __import__("threading").Lock()

# Config
DOC_IDS_FILE = "document_ids_all.txt"
API_BASE_URL = os.environ.get("SECTION301_API_BASE", "https://api.regulations.gov/v4")
DOWNLOAD_BASE = os.environ.get("SECTION301_DOWNLOAD_BASE", "https://downloads.regulations.gov")
API_KEY = os.environ.get("SECTION301_API_KEY", "dwLtc83iWikYr6cF4EDnNAj9TnfWioAHSglZgfSb")
MAX_WORKERS = 2  # Lower to avoid 429 rate limit (regulations.gov ~1000/hr)
API_RATE_DELAY = 4.0  # Seconds between API calls (1000/hr ≈ 1 req/3.6 sec)
OUTPUT_DIR = "section301_list2_output"
# When set: load doc list from Excel, SKIP rows with extraction_method containing "Form Fields" (keep as-is)
EXISTING_EXCEL_PATH = os.environ.get("SECTION301_EXISTING_EXCEL", "section301_list2_output/section301_list2_20260306_215829_UPDATED_REPROCESSED_WITHDRAWN_FILLED_REPROCESSED_REPROCESSED.xlsx")
os.makedirs(OUTPUT_DIR, exist_ok=True)

def _get_all_pdf_urls(doc_id):
    """Get PDF URLs from API + fallback. Retries on 429 with visible backoff.
    Returns (url_titles, api_error). api_error is None if we got 200; else e.g. 'API rate limited (429)'.
    Fallbacks are ONLY added when API returned 200 with no non-Re_ PDFs - never when API failed."""
    global _rate_limit_count
    headers = {"X-Api-Key": API_KEY}
    result = []
    seen = set()
    api_titles = []
    url = f"{API_BASE_URL}/documents/{doc_id}/attachments"
    api_error = None
    for attempt in range(4):
        try:
            time.sleep(API_RATE_DELAY)
            r = requests.get(url, headers=headers, timeout=30)
            if r.status_code == 429:
                wait = (attempt + 1) * 30
                with _rate_limit_lock:
                    _rate_limit_count += 1
                print(f"\n  *** RATE LIMITED (429) for {doc_id} - waiting {wait}s before retry {attempt+1}/4 ***", flush=True)
                time.sleep(wait)
                api_error = "API rate limited (429)"
                continue
            if r.status_code == 200:
                api_error = None
                for att in r.json().get("data", []):
                    title = att.get("attributes", {}).get("title", "") or ""
                    if title:
                        api_titles.append(title)
                    for f in (att.get("attributes", {}).get("fileFormats") or []):
                        if isinstance(f, dict) and f.get("format") == "pdf" and f.get("fileUrl"):
                            u = f["fileUrl"]
                            if u not in seen:
                                result.append((u, title))
                                seen.add(u)
                            break
                break
            # Non-200, non-429
            api_error = f"API error {r.status_code}"
            print(f"\n  *** API {r.status_code} for {doc_id} ***", flush=True)
            if attempt < 3:
                time.sleep(5)
        except Exception as e:
            api_error = f"API error: {str(e)[:50]}"
            print(f"\n  *** API failed for {doc_id}: {api_error} ***", flush=True)
            if attempt < 3:
                time.sleep(5)
    # Fallback: ONLY when API returned 200 successfully and had no non-Re_ PDFs
    if api_error is None:
        non_re_from_api = [(u, t) for u, t in result if not (str(t or "").strip().lower().startswith("re_"))]
        if not non_re_from_api:
            for n in range(1, 10):
                u = f"{DOWNLOAD_BASE}/{doc_id}/attachment_{n}.pdf"
                if u not in seen:
                    result.append((u, ""))
                    seen.add(u)
    return result, api_error


def process_one_doc(doc_id):
    """Process a single document: get ALL PDF URLs, download each, run detection, pick best."""
    result = {
        "documentId": doc_id,
        "Document URL": f"https://www.regulations.gov/document/{doc_id}",
        "attachment_url": "",
        "attachment_title": "",
        "detection_result": "",
        "selection_reason": "No Section 301 form",
        "best_score": 0,
        "status": "pdf not present",
        "total_attachments_checked": 0,
    }
    try:
        url_titles, api_error = _get_all_pdf_urls(doc_id)
        if api_error:
            result["status"] = "API error"
            result["selection_reason"] = api_error
            result["pdf_check_details"] = [("API", False, api_error)]
            return result
        # Skip Re_ attachments (USTR response letters)
        url_titles = [(u, t) for u, t in url_titles if not (str(t or "").strip().lower().startswith("re_"))]
        if not url_titles:
            result["selection_reason"] = "No non-Re_ PDF attachments"
            return result
        candidates = []
        attachments_checked = 0
        pdfs_fetched_ok = 0  # Successfully downloaded and analyzed (no fetch error)
        pdf_check_details = []  # [(name, is_section301, msg)]
        # Download and check EACH PDF - detect_section_301_from_url fetches + runs pattern detection
        for j, (pdf_url, title) in enumerate(url_titles, 1):
            try:
                if not pdf_url:
                    continue
                attachments_checked += 1
                # Derive short name: API title or "attachment_N.pdf"
                pdf_name = (title or "").strip() or pdf_url.split("/")[-1] or f"attachment_{j}.pdf"
                if len(pdf_name) > 50:
                    pdf_name = pdf_name[:47] + "..."
                # Download PDF and run detection - retry on 429
                time.sleep(0.5)
                out = None
                for pdf_attempt in range(3):
                    pdf_headers = {
                        "X-Api-Key": API_KEY,
                        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                        "Referer": "https://www.regulations.gov/",
                    }
                    out = section301_detection.detect_section_301_from_url(
                        pdf_url, f"att_{j}", headers=pdf_headers
                    )
                    msg = out[1] if out else ""
                    if "429" in msg or "Too Many" in msg:
                        wait = 15 * (pdf_attempt + 1)
                        with _rate_limit_lock:
                            _rate_limit_count += 1
                        print(f"\n  *** RATE LIMITED (429) on PDF fetch for {doc_id} - waiting {wait}s ***", flush=True)
                        time.sleep(wait)
                    else:
                        break
                if out and ("429" in (out[1] or "")):
                    pdf_check_details.append((pdf_name, False, "Rate limited (429)"))
                    continue
                is_form = out[0]
                msg = out[1]
                score = out[2]
                if msg and not msg.strip().lower().startswith("error:"):
                    pdfs_fetched_ok += 1
                extraction_method = out[3] if len(out) > 3 else None
                is_section301 = is_form and "EXCLUDED:" not in msg
                pdf_check_details.append((pdf_name, is_section301, msg))
                if "EXCLUDED:" in msg:
                    pass
                elif is_form:
                    candidates.append({
                        "attachment_url": pdf_url,
                        "attachment_title": title or f"Attachment_{j}",
                        "detection_result": msg,
                        "score": score,
                        "extraction_method": extraction_method or "Text",
                    })
            except Exception as e:
                pdf_name = (title or "").strip() or pdf_url.split("/")[-1] or f"attachment_{j}.pdf"
                if len(pdf_name) > 50:
                    pdf_name = pdf_name[:47] + "..."
                pdf_check_details.append((pdf_name, False, f"Error: {str(e)[:30]}"))
                continue
        if candidates:
            best = max(candidates, key=lambda x: x["score"])
            ext_method = best.get("extraction_method", "Text")
            # Scanned = PDF found but no Form Fields (Scanned or Text extraction)
            status_val = "Scanned" if ext_method in ("Scanned", "Text") else "Success"
            result.update({
                "attachment_url": best["attachment_url"],
                "attachment_title": best["attachment_title"],
                "detection_result": best["detection_result"],
                "best_score": best["score"],
                "selection_reason": f"BEST (score={best['score']})",
                "status": status_val,
                "extraction_method": ext_method,
            })
        elif attachments_checked == 0:
            result["selection_reason"] = "No PDF attachments"
        elif pdfs_fetched_ok > 0:
            result["status"] = "pdf not present"
            result["selection_reason"] = "Checked PDFs, none matched"
        else:
            result["selection_reason"] = "All PDFs failed to fetch (404/error)"
        result["total_attachments_checked"] = attachments_checked
        result["pdf_check_details"] = pdf_check_details
    except Exception as e:
        result["selection_reason"] = f"Error: {str(e)[:50]}"
    return result


def _to_complete11k_row(our_result, existing_row=None):
    """Build Complete-11k schema row from our detection result."""
    did = our_result.get("documentId", "")
    row = {
        "documentId": did,
        "Document URL": our_result.get("Document URL", f"https://www.regulations.gov/document/{did}"),
        "Posted Date": "",
        "Last Modified Date": "",
        "attachment_url": our_result.get("attachment_url", ""),
        "USTR Response Date": "",
        "Withdrawn": None,
        "q1_bci_status": "", "q2_product_description": "", "q3_htsus": "", "q4_requestor_name": "",
        "q4_organization": "", "q4_representative": "", "q5_relationship": "", "q6_attachments": "",
        "q7_attachment_bci": "", "q8_us_sources": "", "q9_third_countries": "",
        "q10_2015_value": "", "q10_2015_quantity": "", "q10_2016_value": "", "q10_2016_quantity": "",
        "q10_2017_value": "", "q10_2017_quantity": "", "q11_supporting_info": "",
        "extraction_method": our_result.get("extraction_method", "Data extraction (rule-based)" if our_result.get("status") == "Success" else ""),
        "pipeline_script": "data_extraction_from_txt",
        "status": our_result.get("status", "pdf not present"),
        "filled": 0,
        "form_match_score": our_result.get("best_score", 0),
        "Notes": (
            "Scanned" if our_result.get("status") == "Scanned"
            else (our_result.get("selection_reason", "") if our_result.get("status") == "API error"
            else "PDF not present" if our_result.get("status") == "pdf not present" or not our_result.get("attachment_url")
            else "")
        ),
    }
    details = our_result.get("pdf_check_details", [])
    if details:
        row["PDF_check_details"] = "\n".join(
            f"{n}: {'✓ Section 301' if ok else '✗ (' + str(msg)[:40] + ')'}" for n, ok, msg in details
        )
    if existing_row:
        for k in ["Posted Date", "Last Modified Date", "USTR Response Date", "Withdrawn"]:
            if k in existing_row and existing_row.get(k) is not None and str(existing_row.get(k)).strip():
                row[k] = existing_row[k]
    return row


def add_summary_sheet(excel_path, results, form_fields_count=0):
    """Add Summary sheet and apply color coding to data rows."""
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill
        wb = load_workbook(excel_path)
        ws = wb.active
        ws.title = "Extracted Data"
        # Apply row-level color coding
        red_fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
        gray_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        status_col = withdrawn_col = notes_col = None
        for col_idx, cell in enumerate(ws[1], start=1):
            v = str(cell.value or "").strip().lower()
            if v == "status": status_col = col_idx
            elif v == "withdrawn": withdrawn_col = col_idx
            elif v == "notes": notes_col = col_idx
        if status_col:
            for row_idx in range(2, ws.max_row + 1):
                w = ws.cell(row=row_idx, column=withdrawn_col).value
                withdrawn = w == 1 or (isinstance(w, (int, float)) and int(w) == 1)
                s = str(ws.cell(row=row_idx, column=status_col).value or "").strip()
                fill = None
                if withdrawn:
                    fill = yellow_fill
                    ws.cell(row=row_idx, column=status_col).value = "withdrawn"
                    if notes_col: ws.cell(row=row_idx, column=notes_col).value = "Withdrawn"
                elif s == "Scanned":
                    fill = red_fill
                    if notes_col: ws.cell(row=row_idx, column=notes_col).value = "Scanned"
                elif s in ("pdf not present", "API error"):
                    fill = gray_fill
                    if notes_col and s == "API error": ws.cell(row=row_idx, column=notes_col).value = "API error"
                if fill:
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill
        summary = wb.create_sheet("Summary", 0)
        summary["A1"] = "Metric"
        summary["B1"] = "Count"
        summary["A1"].font = Font(bold=True)
        summary["B1"].font = Font(bold=True)
        total = len(results)
        forms_found = sum(1 for r in results if r.get("status") == "Success")
        form_fields_found = sum(1 for r in results if "Form Fields" in str(r.get("extraction_method", "")))
        pdf_not_present = sum(1 for r in results if r.get("status") == "pdf not present")
        scanned_found = sum(1 for r in results if r.get("status") == "Scanned")
        api_errors = sum(1 for r in results if r.get("status") == "API error")
        withdrawn_count = sum(1 for r in results if r.get("Withdrawn") == 1)
        summary["A2"] = "Total Documents"
        summary["B2"] = total
        summary["A3"] = "Forms Found (Success)"
        summary["B3"] = forms_found
        summary["A4"] = "Form Fields (fillable, detected)"
        summary["B4"] = form_fields_found
        summary["A5"] = "Scanned (red) - PDF found, no Form Fields"
        summary["B5"] = scanned_found
        summary["A6"] = "Form Fields (skipped, kept from existing)"
        summary["B6"] = form_fields_count
        summary["A7"] = "PDF not present (grey)"
        summary["B7"] = pdf_not_present
        summary["A8"] = "Withdrawn (yellow)"
        summary["B8"] = withdrawn_count
        summary["A9"] = "API error (rate limited / failed)"
        summary["B9"] = api_errors
        summary.cell(row=5, column=1).fill = red_fill
        summary.cell(row=5, column=2).fill = red_fill
        summary.cell(row=7, column=1).fill = gray_fill
        summary.cell(row=7, column=2).fill = gray_fill
        summary.cell(row=8, column=1).fill = yellow_fill
        summary.cell(row=8, column=2).fill = yellow_fill
        summary.column_dimensions["A"].width = 35
        summary.column_dimensions["B"].width = 15
        wb.save(excel_path)
        wb.close()
    except Exception as e:
        print(f"Warning: Could not add summary: {e}")


def main():
    print("=" * 60)
    print("Section 301 Data Extraction (GitHub data-extraction logic, no LLM)")
    print("=" * 60)
    print(f"API_BASE_URL: {API_BASE_URL}")
    print(f"DOWNLOAD_BASE (fallback): {DOWNLOAD_BASE}")
    print(f"MAX_WORKERS: {MAX_WORKERS}")
    print(f"EXISTING_EXCEL_PATH: {EXISTING_EXCEL_PATH}")

    doc_list = []
    form_fields_ids = set()
    existing_rows = {}

    # Load from existing Excel when set - skip Form Fields, process rest
    if EXISTING_EXCEL_PATH and os.path.isfile(EXISTING_EXCEL_PATH):
        try:
            df_ex = pd.read_excel(EXISTING_EXCEL_PATH, sheet_name="Extracted Data")
        except Exception:
            try:
                df_ex = pd.read_excel(EXISTING_EXCEL_PATH)
            except Exception:
                df_ex = None
        if df_ex is not None and "documentId" in df_ex.columns:
            seen = set()
            for _, row in df_ex.iterrows():
                did = str(row.get("documentId", ""))
                if did and did not in seen:
                    seen.add(did)
                    doc_list.append({"documentId": did})
            if "extraction_method" in df_ex.columns:
                mask = df_ex["extraction_method"].astype(str).str.contains("Form Fields", na=False)
                form_fields_ids = set(df_ex.loc[mask, "documentId"].astype(str).tolist())
            for _, row in df_ex.iterrows():
                existing_rows[str(row.get("documentId", ""))] = row.to_dict()
            print(f"Loaded existing: {EXISTING_EXCEL_PATH}")
            print(f"Total docs: {len(doc_list)}, Form Fields (SKIP): {len(form_fields_ids)}")
        else:
            print(f"Warning: existing Excel missing or no documentId column")
    if not doc_list and os.path.isfile(DOC_IDS_FILE):
        with open(DOC_IDS_FILE) as f:
            ids = [l.strip() for l in f if l.strip() and not l.startswith("#")]
        doc_list = [{"documentId": did} for did in ids]
        print(f"Loaded {len(doc_list)} IDs from {DOC_IDS_FILE}")
    if not doc_list:
        print("Error: No documents to process. Set EXISTING_EXCEL_PATH or DOC_IDS_FILE.")
        return

    doc_list_to_process = [d for d in doc_list if str(d["documentId"]) not in form_fields_ids]
    print(f"\nProcessing {len(doc_list_to_process)} docs (skipping {len(form_fields_ids)} Form Fields)...")
    print("PDF URLs = API attachments (non-Re_) + fallback (attachment_1..9.pdf) when API has non-Re_")
    print("For each doc: download each PDF -> run detection -> pick best. ✓ = Section 301 form, ✗ = not")

    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as ex:
        futures = {ex.submit(process_one_doc, d["documentId"]): d["documentId"] for d in doc_list_to_process}
        for i, future in enumerate(as_completed(futures)):
            res = future.result()
            results.append(res)
            status = res.get("status", "?")
            score = res.get("best_score", 0)
            n_checked = res.get("total_attachments_checked", "?")
            details = res.get("pdf_check_details", [])
            print(f"  {res.get('documentId')}: {status} (score={score}, PDFs={n_checked})", flush=True)
            if details:
                for pdf_name, ok, msg in details:
                    mark = "✓ Section 301" if ok else f"✗ ({msg})"
                    print(f"      - {pdf_name}: {mark}")
            if (i + 1) % 50 == 0:
                print(f"  Done {i + 1}/{len(doc_list_to_process)}")

    # Merge: Form Fields (from existing) + our processed results
    results_by_id = {str(r.get("documentId")): r for r in results}
    merged = []
    for d in doc_list:
        did = str(d["documentId"])
        if did in form_fields_ids and did in existing_rows:
            merged.append(existing_rows[did])
        elif did in results_by_id:
            merged.append(_to_complete11k_row(results_by_id[did], existing_rows.get(did)))
        else:
            merged.append(_to_complete11k_row({"documentId": did, "status": "pdf not present", "best_score": 0}, existing_rows.get(did)))

    col_order = [
        "documentId", "Document URL", "Posted Date", "Last Modified Date",
        "attachment_url", "USTR Response Date", "Withdrawn",
        "q1_bci_status", "q2_product_description", "q3_htsus", "q4_requestor_name", "q4_organization",
        "q4_representative", "q5_relationship", "q6_attachments", "q7_attachment_bci",
        "q8_us_sources", "q9_third_countries",
        "q10_2015_value", "q10_2015_quantity", "q10_2016_value", "q10_2016_quantity",
        "q10_2017_value", "q10_2017_quantity", "q11_supporting_info",
        "extraction_method", "pipeline_script", "status", "filled", "form_match_score", "Notes",
        "PDF_check_details",
    ]
    df = pd.DataFrame(merged)
    df = df[[c for c in col_order if c in df.columns]]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_file = os.path.join(OUTPUT_DIR, f"section301_data_extraction_{ts}.xlsx")
    df.to_excel(out_file, index=False)
    add_summary_sheet(out_file, merged, form_fields_count=len(form_fields_ids))

    forms = sum(1 for r in merged if r.get("status") == "Success")
    print(f"\nOutput: {out_file}")
    print(f"Total: {len(merged)} (Form Fields kept: {len(form_fields_ids)}, Re-processed: {len(doc_list_to_process)})")
    print(f"Forms found (Success): {forms}")
    if _rate_limit_count > 0:
        print(f"Rate limits (429) hit: {_rate_limit_count} times")
    print("=" * 60)


if __name__ == "__main__":
    main()
