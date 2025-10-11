import pandas as pd
import requests
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import json
import os

# Configuration
API_KEY = "rWRXjXvflW9STOTgc8JKH9kFkkyY7j5qR3GcWba6"
INPUT_FILE = "section301_enhanced_1001_1100_20251011_110612.xlsx"
OUTPUT_FILE = "301-1000IDs_with_date_+_withdrawn_col.xlsx"
PROGRESS_FILE = "withdrawn_progress.json"

# API headers
headers = {
    "X-Api-Key": API_KEY
}

def get_document_details(document_id, max_retries=5):
    """Fetch document details from the API with retry logic"""
    url = f"https://api.regulations.gov/v4/documents/{document_id}"
    
    for attempt in range(max_retries):
        try:
            response = requests.get(url, headers=headers, timeout=30)
            
            # Handle rate limiting
            if response.status_code == 429:
                wait_time = 60 * (attempt + 1)  # Exponential backoff: 60s, 120s, 180s...
                print(f"  Rate limited! Waiting {wait_time} seconds before retry...")
                time.sleep(wait_time)
                continue
            
            response.raise_for_status()
            data = response.json()
            withdrawn = data.get('data', {}).get('attributes', {}).get('withdrawn', False)
            return withdrawn
            
        except requests.exceptions.HTTPError as e:
            if attempt < max_retries - 1:
                wait_time = 30 * (attempt + 1)
                print(f"  HTTP Error: {e}. Retrying in {wait_time} seconds...")
                time.sleep(wait_time)
            else:
                print(f"  Error fetching details for {document_id}: {e}")
                return None
        except Exception as e:
            print(f"  Error fetching details for {document_id}: {e}")
            return None
    
    return None

def process_withdrawn_status(df):
    """Process all documents and get withdrawn status"""
    # Load previous progress if exists
    progress = {}
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'r') as f:
            progress = json.load(f)
        print(f"Loaded progress: {len(progress)} documents already processed")
    
    withdrawn_list = []
    
    for idx, row in df.iterrows():
        doc_id = row['documentId']
        
        # Check if already processed
        if doc_id in progress:
            print(f"Processing {idx+1}/{len(df)}: {doc_id} (from cache)")
            withdrawn_list.append(progress[doc_id])
            continue
        
        print(f"Processing {idx+1}/{len(df)}: {doc_id}")
        
        withdrawn = get_document_details(doc_id)
        
        if withdrawn is None:
            print(f"  Could not fetch withdrawn status")
            withdrawn_list.append(None)
        elif withdrawn:
            print(f"  Withdrawn: YES (1)")
            withdrawn_list.append(1)
        else:
            print(f"  Withdrawn: NO (0)")
            withdrawn_list.append(0)
        
        # Save progress
        progress[doc_id] = withdrawn_list[-1]
        with open(PROGRESS_FILE, 'w') as f:
            json.dump(progress, f)
        
        # Be nice to the API - increased delay to avoid rate limiting
        time.sleep(2)
    
    return withdrawn_list

def apply_red_formatting(excel_file, column_name):
    """Apply red background to cells with value 1 in the specified column"""
    wb = load_workbook(excel_file)
    ws = wb.active
    
    # Find the column index for the withdrawn column
    header_row = 1
    withdrawn_col_idx = None
    
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value == column_name:
            withdrawn_col_idx = col_idx
            break
    
    if withdrawn_col_idx is None:
        print(f"Warning: Could not find column '{column_name}'")
        return
    
    # Define red fill
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Apply red background to cells with value 1
    for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        cell = ws.cell(row=row_idx, column=withdrawn_col_idx)
        if cell.value == 1:
            cell.fill = red_fill
    
    wb.save(excel_file)
    print(f"\nRed formatting applied to {column_name} column")

def main():
    print("Loading Excel file...")
    df = pd.read_excel(INPUT_FILE)
    
    print(f"Found {len(df)} documents to process")
    print(f"Current columns: {list(df.columns)}")
    
    # Process documents and get withdrawn status
    print("\nFetching withdrawn status for all documents...")
    withdrawn_status = process_withdrawn_status(df)
    
    # Add the new column
    df['Withdrawn'] = withdrawn_status
    
    # Report statistics
    withdrawn_count = sum(1 for w in withdrawn_status if w == 1)
    not_withdrawn_count = sum(1 for w in withdrawn_status if w == 0)
    unknown_count = sum(1 for w in withdrawn_status if w is None)
    
    print(f"\n{'='*60}")
    print(f"SUMMARY")
    print(f"{'='*60}")
    print(f"Total documents: {len(df)}")
    print(f"Withdrawn (1): {withdrawn_count}")
    print(f"Not withdrawn (0): {not_withdrawn_count}")
    print(f"Unknown: {unknown_count}")
    
    # Save to Excel file
    print(f"\nSaving results to {OUTPUT_FILE}...")
    df.to_excel(OUTPUT_FILE, index=False)
    
    # Apply red formatting to withdrawn cells
    print("Applying red background formatting to withdrawn documents...")
    apply_red_formatting(OUTPUT_FILE, 'Withdrawn')
    
    print(f"\nDone! Results saved to {OUTPUT_FILE}")
    
    # Clean up progress file
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)
        print(f"Progress file removed")
    
    # Show sample of results
    print("\nSample of results:")
    sample = df[['documentId', 'USTR Response Date', 'Withdrawn']].head(15)
    print(sample.to_string())
    
    # Show withdrawn documents
    if withdrawn_count > 0:
        print("\n\nWithdrawn documents:")
        withdrawn_docs = df[df['Withdrawn'] == 1][['documentId', 'Requester', 'Withdrawn']]
        print(withdrawn_docs.to_string())

if __name__ == "__main__":
    main()